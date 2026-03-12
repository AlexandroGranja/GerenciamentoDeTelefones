from __future__ import annotations

import io
import json
import os
import secrets
import base64
import subprocess
import sys
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st

# Cookie manager para login persistente (inicializado em main, após set_page_config)
HAS_COOKIES = False
_cookies = None

def _init_cookies():
    """Inicializa o gerenciador de cookies. Chamar após set_page_config."""
    global _cookies, HAS_COOKIES
    if _cookies is not None:
        return
    try:
        from streamlit_cookies_manager import EncryptedCookieManager
        _cookies = EncryptedCookieManager(
            prefix="planilhas_tel/",
            password=os.environ.get("COOKIES_PASSWORD", "planilhas_telefones_secret_2024"),
        )
        HAS_COOKIES = True
    except ImportError:
        HAS_COOKIES = False
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Suporte a banco de dados (opcional)
try:
    from src.db.repository import (
        load_linhas, has_data, get_db_path, save_linhas,
        init_db, verificar_login, criar_usuario, listar_usuarios, excluir_usuario, tem_usuarios,
        obter_usuario_por_username, criar_sessao, validar_sessao, encerrar_sessao,
        registrar_auditoria, listar_auditoria,
    )
    HAS_DB = True
except ImportError:
    HAS_DB = False
    init_db = verificar_login = criar_usuario = listar_usuarios = excluir_usuario = tem_usuarios = None
    obter_usuario_por_username = criar_sessao = validar_sessao = encerrar_sessao = None
    registrar_auditoria = listar_auditoria = None


from src.core.config import (
    DEFAULT_FILE, DEFAULT_FILE_COMPLETO, RELACAO_FILE, RULES_FILE,
    PLANILHAS_DIR, DOC_DIR,
)
ABAS_ALIMENTO = ["Nova Prosper"]
ABAS_MEDICAMENTO = ["Prosper Norte", "Prosper Sul"]
ABAS_FOCO = ["Prosper Norte", "Prosper Sul", "Nova Prosper", "Promotores", "Internos", "Troca de Aparelho", "Devolução Manutenção", "Roubo-Perda"]
ABAS_PROMOTORES = ["Promotores"]
ABAS_INTERNOS = ["Internos"]
ABAS_MANUTENCAO = ["Troca de Aparelho", "Devolução Manutenção", "Devolucao Manutencao"]
ABAS_ROUBO_PERDA = ["Roubo-Perda", "Roubo e Perda"]
EQUIPES_PROMOTORES = ["Promotores"]
EQUIPES_INTERNOS = ["Internos"]
EQUIPES_MANUTENCAO = ["Manutenção"]
EQUIPES_ROUBO_PERDA = ["Roubo e Perda"]
EQUIPES_ALIMENTO = [
    "Gerentes do Alimento",
    "Consumo Baixada",
    "Consumo Oeste",
    "Consumo Zona Norte",
    "Consumo Niteroi",
    "Equipe Especial",
    "Gerente Senior",
]
EQUIPES_MEDICAMENTO = [
    "Gerentes do Medicamento",
    "Prosper Norte",
    "Prosper Sul",
]
GESTORES_MEDICAMENTO = {
    "Prosper Norte": "Priscila Rangel Manhães",
    "Prosper Sul": "Gustavo Luis Dias De Armada",
}
DEFAULT_COLUMNS = [
    "Codigo",
    "Nome",
    "Nome de Guerra",
    "Equipe",
    "Linha",
    "E-mail",
    "Gerenciamento",
    "Data da Troca",
    "Data Retorno",
    "Data Ocorrência",
    "Data Solicitação TBS",
    "Motivo",
    "Observação",
    "IMEI A",
    "IMEI B",
    "Marca",
    "CHIP",
    "Aparelho",
    "Modelo",
    "Setor",
    "Cargo",
    "Desconto",
    "Perfil",
    "Empresa",
    "Ativo",
    "Numero de Serie",
    "Patrimonio",
    "Operadora",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = "".join(
        ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn"
    )
    return text.lower()


def title_case_safe(value: str) -> str:
    tokens = [t for t in value.replace("/", " / ").split(" ") if t != ""]
    if not tokens:
        return "Sem Equipe"
    out: list[str] = []
    for token in tokens:
        if token == "/":
            out.append(token)
        elif token.isupper() and len(token) <= 4:
            out.append(token)
        else:
            out.append(token.capitalize())
    return " ".join(out).replace(" / ", "/")


def normalize_team_key(value: Any) -> str:
    text = normalize_text(value)
    text = " ".join(text.split())
    return text


def digits_only(value: Any) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value) if ch.isdigit())


def is_valid_phone(value: Any) -> bool:
    digits = digits_only(value)
    return 10 <= len(digits) <= 13


def find_header_row(ws) -> tuple[int | None, int | None]:
    max_r = min(80, ws.max_row)
    max_c = min(80, ws.max_column)
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            if normalize_text(ws.cell(row, col).value) == "linha":
                return row, col
    return None, None


def collect_headers(ws, header_row: int, start_col: int, limit: int = 40) -> list[str]:
    headers: list[str] = []
    empty_streak = 0
    max_c = min(ws.max_column, start_col + limit)
    for col in range(1, max_c + 1):
        value = ws.cell(header_row, col).value
        text = str(value).strip() if value is not None else ""
        if text == "":
            headers.append(f"col_{col}")
            empty_streak += 1
        else:
            headers.append(text)
            empty_streak = 0
        if col >= start_col and empty_streak >= 8:
            break
    return headers


def map_known_column(column_name: str) -> str:
    c = normalize_text(column_name)
    mapping = {
        "codigo": "Codigo",
        "codigos": "Codigo",
        "asset": "Codigo",
        "asset tag": "Codigo",
        "nome": "Nome",
        "nomes": "Nome",
        "nome colaborador": "Nome",
        "nome de guerra": "Nome de Guerra",
        "equipe": "Equipe",
        "local": "Localidade",
        "localidade": "Localidade",
        "data da troca": "Data da Troca",
        "data troca": "Data da Troca",
        "data ocorrencia": "Data Ocorrência",
        "data ocorrência": "Data Ocorrência",
        "data retorno": "Data Retorno",
        "data de retorno": "Data Retorno",
        "data solicitacao tbs": "Data Solicitação TBS",
        "data solicitação tbs": "Data Solicitação TBS",
        "linha": "Linha",
        "email": "E-mail",
        "e-mail": "E-mail",
        "gerenciamento": "Gerenciamento",
        "bloqueio": "Gerenciamento",
        "imei": "IMEI A",
        "imei a": "IMEI A",
        "imei2": "IMEI B",
        "imei b": "IMEI B",
        "chip": "CHIP",
        "marca": "Marca",
        "aparelho": "Aparelho",
        "modelo": "Modelo",
        "setor": "Setor",
        "cargo": "Cargo",
        "desconto": "Desconto",
        "perfil": "Perfil",
        "empresa": "Empresa",
        "ativo": "Ativo",
        "ativos": "Ativo",
        "numero de serie": "Numero de Serie",
        "patrimonio": "Patrimonio",
        "motivo": "Motivo",
        "obs": "Observação",
        "ns": "Numero de Serie",
        "operadora": "Operadora",
        "s/n": "Numero de Serie",
        "s n": "Numero de Serie",
        "nº serie": "Numero de Serie",
        # Prosper Norte usa nome do gerente como cabeçalho da coluna código
        "fernando goncalves de mello": "Codigo",
        "goncalves d": "Codigo",
        "goncalves": "Codigo",
    }
    return mapping.get(c, column_name)


@st.cache_data(show_spinner=False)
def load_active_lines(file_path: str) -> pd.DataFrame:
    workbook = load_workbook(file_path, data_only=True)
    rows: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        header_row, line_col = find_header_row(ws)

        if header_row is None or line_col is None:
            # Fallback for sheets without explicit header row.
            headers = DEFAULT_COLUMNS.copy()
            first_data_row = 1
            max_col = len(headers)
        else:
            headers = collect_headers(ws, header_row, line_col)
            headers = [map_known_column(h) for h in headers]
            first_data_row = header_row + 1
            max_col = len(headers)

        for row_idx in range(first_data_row, ws.max_row + 1):
            line_value = ws.cell(row_idx, 4).value if header_row is None else ws.cell(row_idx, line_col).value
            if not is_valid_phone(line_value):
                continue

            row_data: dict[str, Any] = {}
            seen_keys: dict[str, int] = {}
            for col_idx in range(1, max_col + 1):
                key = headers[col_idx - 1] if col_idx - 1 < len(headers) else f"col_{col_idx}"
                seen_keys[key] = seen_keys.get(key, 0) + 1
                # Duas colunas "IMEI" na planilha de Roubo/Perda: preservar ambas.
                if key == "IMEI A" and seen_keys[key] == 2:
                    key = "IMEI B"
                row_data[key] = ws.cell(row_idx, col_idx).value

            row_data["Linha"] = digits_only(line_value)
            row_data["Aba"] = sheet_name
            rows.append(row_data)

    if not rows:
        return pd.DataFrame(columns=DEFAULT_COLUMNS + ["Aba"])

    df = pd.DataFrame(rows)

    # Guarantee key columns.
    for col in DEFAULT_COLUMNS + ["Aba"]:
        if col not in df.columns:
            df[col] = ""

    ordered = DEFAULT_COLUMNS + ["Aba"]
    other_cols = [c for c in df.columns if c not in ordered]
    df = df[ordered + other_cols]

    df["Equipe"] = df["Equipe"].fillna("Sem Equipe").astype(str).str.strip()
    df.loc[df["Equipe"] == "", "Equipe"] = "Sem Equipe"
    df["Nome"] = df["Nome"].fillna("").astype(str).str.strip()
    df["Aparelho"] = df["Aparelho"].fillna("").astype(str).str.strip()
    df["Modelo"] = df["Modelo"].fillna("").astype(str).str.strip()
    df["Codigo"] = df["Codigo"].fillna("").astype(str).str.strip()

    return df


@st.cache_data(show_spinner=False)
def load_active_lines_set(relacao_path: str) -> frozenset[str]:
    """Retorna o conjunto de linhas ativas (na Relação ListaAtual)."""
    try:
        wb = load_workbook(relacao_path, data_only=True)
        if "ListaAtual" not in wb.sheetnames:
            return frozenset()
        ws = wb["ListaAtual"]
        ativas: set[str] = set()
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 4).value
            d = digits_only(v)
            if 10 <= len(d) <= 13:
                ativas.add(d)
        return frozenset(ativas)
    except Exception:
        return frozenset()


def load_inactive_lines(file_path: str, relacao_path: str) -> pd.DataFrame:
    """Carrega linhas de Telefones que NÃO estão na Relação (desativadas). Apenas abas em ABAS_FOCO."""
    df_full = load_active_lines(file_path)
    ativas = load_active_lines_set(relacao_path)
    abas_norm = {a.strip().lower() for a in ABAS_FOCO}
    mask_aba = df_full["Aba"].fillna("").astype(str).str.strip().str.lower().isin(abas_norm)
    mask_inativa = ~df_full["Linha"].isin(ativas)
    df = df_full[mask_aba & mask_inativa].copy()
    return df


def detect_tipo_equipe(equipe: str, aba: str) -> str:
    e = normalize_text(equipe)
    a = normalize_text(aba)
    if "intern" in e or "intern" in a:
        return "Interna"
    return "Externa"


def detect_localidade(equipe: str) -> str:
    e = str(equipe).strip()
    if e == "":
        return ""
    markers = ["gerencia", "gerente", "supervisor", "promotor", "diretoria", "internos"]
    low = normalize_text(e)
    if any(marker in low for marker in markers):
        return ""
    return title_case_safe(e)


def _build_alimento_map(rules_path: Path) -> dict[str, dict[str, str]]:
    """Monta mapeamento localidade -> equipe, gestor, supervisor para Alimento (equipes_alimento.csv)."""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    out: dict[str, dict[str, str]] = {}
    team_defaults: dict[str, dict[str, str]] = {}

    if not ali_path.exists():
        return out

    try:
        ref = pd.read_csv(ali_path, dtype=str).fillna("")
        for _, r in ref.iterrows():
            er = str(r.get("equipe_real", "")).strip()
            loc = str(r.get("localidade", "")).strip()
            gerente = str(r.get("gerente", "")).strip()
            supervisor = str(r.get("supervisor", "")).strip()
            if loc:
                key = normalize_team_key(loc)
                out[key] = {"equipe": er, "gestor": gerente, "supervisor": supervisor}
            else:
                eq_key = normalize_team_key(er)
                team_defaults[eq_key] = {"equipe": er, "gestor": gerente, "supervisor": supervisor}
        for eq in ["consumo baixada", "consumo oeste", "consumo zona norte", "consumo niteroi"]:
            out[eq] = team_defaults.get(eq, {"equipe": title_case_safe(eq), "gestor": "Marcelo Neves", "supervisor": ""})
        out["especial consumo"] = {"equipe": "Equipe Especial", "gestor": "Marco Antonio Neves Suzart", "supervisor": "Ricardo Cascao"}
        for loc, eq, sup in [
            ("alcantara", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("cabucu", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("piabetá", "Consumo Baixada", ""),
            ("piabeta", "Consumo Baixada", ""),
            ("realengo", "Consumo Oeste", "Marcelo Martins Da Costa"),
            ("barra da tijuca", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("recreio", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("botafogo", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("taquara", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("anchieta", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("lote xv", "Consumo Baixada", ""),
            ("niteroi i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("niterói i", "Consumo Niteroi", "Fabio Antonio Rosa Magalhaes"),
            ("conveniencia farma 12", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("conveniencia farma 20", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("key account", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("supervisora senior", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("gerencia varejo ii", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("consumo zona sul", "Consumo Zona Norte", "Paulo Roberto Ferreira Chaves"),
            ("itaborai", "Consumo Baixada", ""),
        ]:
            out[normalize_team_key(loc)] = {"equipe": eq, "gestor": "Marcelo Neves", "supervisor": sup}
        for k in ["rede especial 1", "rede especial 2", "rede especial 3", "rede especial 4", "rota especial 1", "rota especial 2",
                  "rota especial 3", "rota especial 4", "impulso 01", "impulso 02", "impulso 03", "impulso 04",
                  "impulso 05", "auto servico 01", "auto servico 02", "auto servico 03", "auto servico 04",
                  "auto servico 05", "auto servico 06", "a s impulso 2", "a s impulso 3", "a s impulso 4",
                  "a s impulso 5"]:
            out[k] = {"equipe": "Equipe Especial", "gestor": "Marco Antonio Neves Suzart", "supervisor": "Ricardo Cascao"}
    except Exception:
        pass
    return out


def detect_grupo_equipe(equipe: str, tipo: str) -> str:
    e = normalize_text(equipe)
    special_tokens = ["especial", "impulso", "auto servico", "auto servico", "rota especial"]
    if any(token in e for token in special_tokens):
        return "Equipe Especial"
    if normalize_text(tipo) == "interna":
        return "Equipe Interna"
    return "Outras Equipes Externas"


def ensure_rules_file(df: pd.DataFrame, rules_path: Path) -> pd.DataFrame:
    if rules_path.exists():
        rules = pd.read_csv(rules_path, dtype=str).fillna("")
        # Backward-compatible migration of rule columns.
        if "equipe_key" not in rules.columns:
            rules["equipe_key"] = rules.get("equipe_origem", "").map(normalize_team_key)
        if "equipe_origem" not in rules.columns:
            rules["equipe_origem"] = ""
        if "equipe_padrao" not in rules.columns:
            rules["equipe_padrao"] = rules["equipe_origem"]
        if "tipo_equipe" not in rules.columns:
            rules["tipo_equipe"] = "Externa"
        if "localidade" not in rules.columns:
            rules["localidade"] = ""
        if "grupo_equipe" not in rules.columns:
            rules["grupo_equipe"] = rules.apply(
                lambda r: detect_grupo_equipe(str(r.get("equipe_padrao", "")), str(r.get("tipo_equipe", ""))),
                axis=1,
            )
        if "gestor" not in rules.columns:
            rules["gestor"] = ""
        if "supervisor" not in rules.columns:
            rules["supervisor"] = ""
        if "segmento" not in rules.columns:
            rules["segmento"] = "Alimento"
        if "eh_equipe" not in rules.columns:
            rules["eh_equipe"] = "True"
        if "equipe_pai" not in rules.columns:
            rules["equipe_pai"] = ""
        rules = rules[
            [
                "equipe_key",
                "equipe_origem",
                "equipe_padrao",
                "grupo_equipe",
                "tipo_equipe",
                "localidade",
                "gestor",
                "supervisor",
                "segmento",
                "eh_equipe",
                "equipe_pai",
            ]
        ]
        rules.to_csv(rules_path, index=False, encoding="utf-8-sig")
        return rules

    base = (
        df[["Equipe", "Aba"]]
        .drop_duplicates()
        .assign(equipe_key=lambda x: x["Equipe"].map(normalize_team_key))
        .sort_values("Equipe")
    )
    base["equipe_origem"] = base["Equipe"]
    base["equipe_padrao"] = base["Equipe"].map(lambda v: title_case_safe(str(v)))
    base["tipo_equipe"] = base.apply(lambda r: detect_tipo_equipe(str(r["Equipe"]), str(r["Aba"])), axis=1)
    base["localidade"] = base["Equipe"].map(detect_localidade)
    base["grupo_equipe"] = base.apply(
        lambda r: detect_grupo_equipe(str(r["equipe_padrao"]), str(r["tipo_equipe"])),
        axis=1,
    )
    base["gestor"] = ""
    base["supervisor"] = ""
    base["segmento"] = "Alimento"
    base["eh_equipe"] = "True"
    base["equipe_pai"] = ""
    rules = base[
        [
            "equipe_key",
            "equipe_origem",
            "equipe_padrao",
            "grupo_equipe",
            "tipo_equipe",
            "localidade",
            "gestor",
            "supervisor",
            "segmento",
            "eh_equipe",
            "equipe_pai",
        ]
    ]
    rules.to_csv(rules_path, index=False, encoding="utf-8-sig")
    return rules


def apply_team_standardization(df: pd.DataFrame, rules_path: Path) -> pd.DataFrame:
    rules = ensure_rules_file(df, rules_path)
    merged = df.copy()
    merged["equipe_key"] = merged["Equipe"].map(normalize_team_key)

    dedup_rules = rules.drop_duplicates(subset=["equipe_key"], keep="first")
    merged = merged.merge(
        dedup_rules[
            [
                "equipe_key",
                "equipe_padrao",
                "grupo_equipe",
                "tipo_equipe",
                "localidade",
                "gestor",
                "supervisor",
                "segmento",
            ]
        ],
        on="equipe_key",
        how="left",
    )

    merged["EquipePadrao"] = merged["equipe_padrao"].fillna("").astype(str).str.strip()
    merged.loc[merged["EquipePadrao"] == "", "EquipePadrao"] = merged["Equipe"].map(
        lambda v: title_case_safe(str(v))
    )

    merged["TipoEquipe"] = merged["tipo_equipe"].fillna("").astype(str).str.strip()
    merged.loc[merged["TipoEquipe"] == "", "TipoEquipe"] = merged.apply(
        lambda r: detect_tipo_equipe(str(r["Equipe"]), str(r["Aba"])),
        axis=1,
    )

    merged["GrupoEquipe"] = merged["grupo_equipe"].fillna("").astype(str).str.strip()
    merged.loc[merged["GrupoEquipe"] == "", "GrupoEquipe"] = merged.apply(
        lambda r: detect_grupo_equipe(str(r["EquipePadrao"]), str(r["TipoEquipe"])),
        axis=1,
    )

    merged["Localidade"] = merged["localidade"].fillna("").astype(str).str.strip()
    merged["Gestor"] = merged["gestor"].fillna("").astype(str).str.strip()
    merged["Supervisor"] = merged["supervisor"].fillna("").astype(str).str.strip()
    merged["Segmento"] = merged["segmento"].fillna("Alimento").astype(str).str.strip()
    merged.loc[merged["Segmento"] == "", "Segmento"] = "Alimento"
    merged.drop(
        columns=[
            "equipe_key",
            "equipe_padrao",
            "grupo_equipe",
            "tipo_equipe",
            "localidade",
            "gestor",
            "supervisor",
            "segmento",
        ],
        inplace=True,
    )

    # Alimento (Nova Prosper): aplicar mesmo padrão de equipes/gerentes que linhas ativas
    mask_ali = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_ALIMENTO]
    )
    if mask_ali.any():
        ali_map = _build_alimento_map(rules_path)
        for idx in merged.index[mask_ali]:
            key = normalize_team_key(merged.at[idx, "Equipe"])
            if key in ali_map:
                merged.at[idx, "EquipePadrao"] = ali_map[key]["equipe"]
                merged.at[idx, "Gestor"] = ali_map[key]["gestor"]
                merged.at[idx, "Supervisor"] = ali_map[key]["supervisor"]
                merged.at[idx, "Localidade"] = merged.at[idx, "Equipe"]
            else:
                # Fallback: Equipe Especial ou Consumo conforme padrão do nome
                low = key
                if any(t in low for t in ["especial", "impulso", "rota especial", "auto servico", "a s impulso"]):
                    merged.at[idx, "EquipePadrao"] = "Equipe Especial"
                    merged.at[idx, "Gestor"] = "Marco Antonio Neves Suzart"
                    merged.at[idx, "Supervisor"] = "Ricardo Cascao"
                else:
                    merged.at[idx, "EquipePadrao"] = "Consumo Zona Norte"
                    merged.at[idx, "Gestor"] = "Marcelo Neves"
                    merged.at[idx, "Supervisor"] = "Paulo Roberto Ferreira Chaves"
                merged.at[idx, "Localidade"] = merged.at[idx, "Equipe"]

    # Medicamento: Prosper Norte/Sul — EquipePadrao = Aba, Gestor por aba. Sem supervisor (não usar regras de Alimento).
    mask_med = merged["Aba"].fillna("").astype(str).str.strip().isin(ABAS_MEDICAMENTO)
    if mask_med.any():
        merged.loc[mask_med, "Segmento"] = "Medicamento"
        merged.loc[mask_med, "EquipePadrao"] = merged.loc[mask_med, "Aba"]
        merged.loc[mask_med, "Localidade"] = merged.loc[mask_med, "Equipe"]
        merged.loc[mask_med, "Supervisor"] = ""
        for aba, gestor in GESTORES_MEDICAMENTO.items():
            merged.loc[mask_med & (merged["Aba"] == aba), "Gestor"] = gestor

    # Promotores (aba) — Segmento Promotores
    mask_prom = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_PROMOTORES]
    )
    if mask_prom.any():
        merged.loc[mask_prom, "Segmento"] = "Promotores"
        merged.loc[mask_prom, "EquipePadrao"] = "Promotores"
        merged.loc[mask_prom, "Localidade"] = merged.loc[mask_prom, "Equipe"]

    # Internos (aba) — Segmento Internos
    mask_int = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_INTERNOS]
    )
    if mask_int.any():
        merged.loc[mask_int, "Segmento"] = "Internos"
        merged.loc[mask_int, "EquipePadrao"] = "Internos"
        mask_loc_vazio = mask_int & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    # Manutenção (abas de troca e devolução)
    mask_manut = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_MANUTENCAO]
    )
    if mask_manut.any():
        merged.loc[mask_manut, "Segmento"] = "Manutenção"
        merged.loc[mask_manut, "EquipePadrao"] = "Manutenção"
        mask_loc_vazio = mask_manut & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    # Roubo e Perda
    mask_roubo = merged["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_ROUBO_PERDA]
    )
    if mask_roubo.any():
        merged.loc[mask_roubo, "Segmento"] = "Roubo e Perda"
        merged.loc[mask_roubo, "EquipePadrao"] = "Roubo e Perda"
        mask_loc_vazio = mask_roubo & (merged["Localidade"].fillna("").astype(str).str.strip() == "")
        merged.loc[mask_loc_vazio, "Localidade"] = merged.loc[mask_loc_vazio, "Equipe"]

    return merged


def apply_filters(
    df: pd.DataFrame,
    grupos: list[str],
    segmentos: list[str],
    teams: list[str],
    tipos: list[str],
    localidades: list[str],
    query: str,
) -> pd.DataFrame:
    result = df.copy()
    if grupos:
        result = result[result["GrupoEquipe"].isin(grupos)]
    if segmentos:
        result = result[result["Segmento"].isin(segmentos)]
    if teams:
        result = result[result["EquipePadrao"].isin(teams)]
    if tipos:
        result = result[result["TipoEquipe"].isin(tipos)]
    if localidades:
        result = result[result["Localidade"].isin(localidades)]

    query = query.strip().lower()
    if query:
        columns = [
            "Linha",
            "Codigo",
            "Nome",
            "Equipe",
            "GrupoEquipe",
            "Segmento",
            "EquipePadrao",
            "TipoEquipe",
            "Localidade",
            "Gestor",
            "Supervisor",
            "Aparelho",
            "Modelo",
            "Aba",
        ]
        mask = pd.Series(False, index=result.index)
        for col in columns:
            mask = mask | result[col].astype(str).str.lower().str.contains(query, na=False)
        result = result[mask]

    return result.sort_values(by=["Segmento", "GrupoEquipe", "EquipePadrao", "Nome", "Linha"], kind="stable")


def non_empty_or_default(value: Any, default: str) -> str:
    text = str(value).strip() if value is not None else ""
    return text if text else default


def normalize_name(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def classify_papel(row: pd.Series) -> str:
    nome = normalize_name(row.get("Nome", ""))
    gestor = normalize_name(row.get("Gestor", ""))
    supervisor = normalize_name(row.get("Supervisor", ""))
    if nome != "" and gestor != "" and nome == gestor:
        return "Gerente"
    if nome != "" and supervisor != "" and nome == supervisor:
        return "Supervisor"
    if "vago" in normalize_text(row.get("Nome", "")):
        return "Vago"
    return "Vendedor"


def build_full_table(df: pd.DataFrame) -> pd.DataFrame:
    table = df.copy()
    table["Papel"] = table.apply(classify_papel, axis=1)
    cols = [
        "Segmento",
        "Data da Troca",
        "Data Retorno",
        "Data Ocorrência",
        "Data Solicitação TBS",
        "GrupoEquipe",
        "EquipePadrao",
        "TipoEquipe",
        "Localidade",
        "Gestor",
        "Supervisor",
        "Papel",
        "Nome de Guerra",
        "Codigo",
        "Nome",
        "Linha",
        "E-mail",
        "Gerenciamento",
        "Motivo",
        "Observação",
        "Marca",
        "Aparelho",
        "Modelo",
        "CHIP",
        "IMEI A",
        "IMEI B",
        "Setor",
        "Cargo",
        "Desconto",
        "Perfil",
        "Empresa",
        "Ativo",
        "Numero de Serie",
        "Patrimonio",
        "Operadora",
        "Aba",
        "Equipe",
    ]
    existing = [c for c in cols if c in table.columns]
    return table[existing].sort_values(
        ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
        kind="stable",
    )


def _get_gerentes_medicamento(rules_path: Path) -> set[str]:
    """Retorna o conjunto de nomes normalizados dos gerentes do Medicamento."""
    out: set[str] = set()
    for nome in GESTORES_MEDICAMENTO.values():
        if nome:
            out.add(normalize_text(nome))
    med_path = rules_path.parent / "equipes_medicamento.csv"
    if med_path.exists():
        try:
            ref = pd.read_csv(med_path, dtype=str).fillna("")
            for _, r in ref.iterrows():
                g = str(r.get("gerente", "")).strip()
                if g:
                    out.add(normalize_text(g))
        except Exception:
            pass
    return out


def _get_gerentes_alimento(rules_path: Path) -> set[str]:
    """Retorna o conjunto de nomes normalizados dos gerentes do Alimento (equipes_alimento.csv)."""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    out: set[str] = set()
    if not ali_path.exists():
        return out
    try:
        ref = pd.read_csv(ali_path, dtype=str).fillna("")
        for _, r in ref.iterrows():
            g = str(r.get("gerente", "")).strip()
            if g:
                out.add(normalize_text(g))
        if "marco antonio neves suzart" not in out:
            out.add("marco antonio neves suzart")
        if "marcelo neves" not in out:
            out.add("marcelo neves")
    except Exception:
        pass
    return out


def _get_supervisor_from_rules(equipe: str, rules_path: Path) -> str:
    """Obtém supervisor da equipe. Medicamento usa equipes_medicamento (geralmente vago)."""
    eq_key = normalize_team_key(equipe)
    if eq_key in ("prosper norte", "prosper sul"):
        med_path = rules_path.parent / "equipes_medicamento.csv"
        if med_path.exists():
            try:
                ref = pd.read_csv(med_path, dtype=str).fillna("")
                for _, r in ref.iterrows():
                    if normalize_team_key(str(r.get("equipe_real", ""))) == eq_key:
                        sup = str(r.get("supervisor", "")).strip()
                        return sup if sup else ""
            except Exception:
                pass
        return ""
    ali_path = rules_path.parent / "equipes_alimento.csv"
    if ali_path.exists():
        try:
            ref = pd.read_csv(ali_path, dtype=str).fillna("")
            fallback_sup = ""
            for _, r in ref.iterrows():
                er = str(r.get("equipe_real", "")).strip()
                loc = str(r.get("localidade", "")).strip()
                if normalize_team_key(er) != eq_key:
                    continue
                sup = str(r.get("supervisor", "")).strip()
                if not sup:
                    continue
                if not loc:
                    return sup
                fallback_sup = sup
            if fallback_sup:
                return fallback_sup
        except Exception:
            pass
    if rules_path.exists():
        try:
            rules = pd.read_csv(rules_path, dtype=str).fillna("")
            for _, r in rules.iterrows():
                if normalize_team_key(str(r.get("equipe_padrao", ""))) == eq_key:
                    sup = str(r.get("supervisor", "")).strip()
                    if sup:
                        return sup
        except Exception:
            pass
    return ""


def _supervisor_display(dfe: pd.DataFrame, eq: str, rules_path: Path) -> str:
    """Nome do supervisor para exibição no cabeçalho. Vago se vazio, senão nome. Fallback em regras."""
    sup_col = dfe["Supervisor"].fillna("").astype(str).str.strip()
    sup = ""
    for v in sup_col:
        if v and v not in ("—", "-", "vago", "Sem Supervisor"):
            sup = v
            break
    if sup:
        return sup
    if eq == "Promotores":
        for _, row in dfe.iterrows():
            loc = normalize_text(str(row.get("Localidade", "") or row.get("Equipe", "")).strip())
            if "supervisor" in loc:
                return non_empty_or_default(row.get("Nome", ""), "—")
    sup_from_rules = _get_supervisor_from_rules(eq, rules_path)
    if sup_from_rules:
        return sup_from_rules
    return "Vago"


def _codigo_supervisor(dfe: pd.DataFrame) -> str:
    """Obtem o codigo do supervisor (linha onde Nome == Supervisor)."""
    sup = dfe["Supervisor"].fillna("").astype(str).str.strip()
    nomes = dfe["Nome"].fillna("").astype(str).str.strip()
    cods = dfe["Codigo"].fillna("").astype(str).str.strip()
    for i, s in enumerate(sup):
        if s and s != "Sem Supervisor":
            nm = nomes.iloc[i]
            if nm and normalize_text(nm) == normalize_text(s):
                return cods.iloc[i] or "—"
    return "—"


def _celula_vago(v: Any) -> str:
    txt = str(v).strip() if v is not None else ""
    if normalize_text(txt) == "vago":
        return '<span style="color:#c00;font-weight:600">VAGO</span>'
    return _escape_html(txt) if txt else ""


def _escape_html(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _render_equipe_tabela(dfe: pd.DataFrame, eq: str) -> str:
    """Tabela com supervisor como 1ª linha e demais vendedores. Cabeçalho com rótulos claros."""
    cod_sup = _codigo_supervisor(dfe)
    nome_sup = non_empty_or_default(dfe["Supervisor"].iloc[0], "—")

    cols_full = [
        "Codigo", "Nome", "Equipe", "Linha", "E-mail", "Gerenciamento",
        "Aparelho", "Modelo", "CHIP", "IMEI A", "IMEI B",
        "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
        "Numero de Serie", "Operadora",
    ]
    labels_full = [
        "Código", "Nome", "Localidade", "Linha", "E-mail", "Gerenciamento",
        "Aparelho", "Modelo", "CHIP", "IMEI A", "IMEI B",
        "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
        "Nº Série", "Operadora",
    ]
    cols = [c for c in cols_full if c in dfe.columns]
    labels = [labels_full[cols_full.index(c)] for c in cols]

    sup_norm = normalize_text(nome_sup)
    row_supervisor = None
    rows_data = []

    def _is_supervisor_row(row: pd.Series) -> bool:
        if sup_norm and normalize_text(str(row.get("Nome", "")).strip()) == sup_norm:
            return True
        if eq == "Promotores":
            loc = normalize_text(str(row.get("Localidade", "") or row.get("Equipe", "")).strip())
            if "supervisor" in loc:
                return True
        return False

    for _, row in dfe.iterrows():
        if _is_supervisor_row(row):
            row_supervisor = row
            continue
        rows_data.append(row)

    html = ['<table class="tbl-equipe">']
    html.append("<thead><tr>")
    for L in labels:
        html.append(f"<th>{L}</th>")
    html.append("</tr></thead><tbody>")

    if row_supervisor is not None:
        display_nome_sup = nome_sup if (nome_sup and nome_sup != "—") else non_empty_or_default(row_supervisor.get("Nome", ""), "—")
        html.append('<tr class="row-supervisor">')
        for c in cols:
            v = row_supervisor.get(c, "")
            s = str(v).strip() if v is not None else ""
            if not s and c == "Codigo":
                s = cod_sup or non_empty_or_default(row_supervisor.get("Codigo", ""), "—")
            elif not s and c == "Nome":
                s = display_nome_sup
            elif not s and c == "Equipe":
                s = eq
            html.append(f'<td><strong>{_escape_html(s)}</strong></td>')
        html.append("</tr>")

    for row in rows_data:
        html.append("<tr>")
        for c in cols:
            v = row.get(c, "")
            s = str(v).strip() if v is not None else ""
            if c in ("Codigo", "Nome") and normalize_text(s) == "vago":
                html.append(f'<td>{_celula_vago(s)}</td>')
            else:
                html.append(f'<td>{_escape_html(s)}</td>')
        html.append("</tr>")
    html.append("</tbody></table>")
    return '<div class="tbl-wrapper">' + "\n".join(html) + '</div>'


def _salvar_sessao_cookie(user: dict) -> None:
    """Cria sessão e salva token no cookie para login persistente."""
    token = secrets.token_hex(32)
    criar_sessao(token, user["username"])
    st.session_state.auth_token = token
    if HAS_COOKIES and _cookies.ready():
        _cookies["auth_token"] = token
        _cookies.save()


def _restaurar_sessao_cookie() -> bool:
    """Tenta restaurar sessão a partir do cookie. Retorna True se restaurou."""
    if not (HAS_COOKIES and HAS_DB and _cookies.ready()):
        return False
    token = _cookies.get("auth_token")
    if not token:
        return False
    user = validar_sessao(token)
    if user:
        st.session_state.authenticated = True
        st.session_state.user = user
        st.session_state.auth_token = token
        return True
    # Token inválido ou expirado
    try:
        del _cookies["auth_token"]
        _cookies.save()
    except Exception:
        pass
    return False


def _render_login_or_first_user() -> bool:
    """
    Renderiza login ou formulário do primeiro usuário.
    Retorna True se o usuário está autenticado (mostrar app principal).
    """
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "user" not in st.session_state:
        st.session_state.user = None
    if "auth_token" not in st.session_state:
        st.session_state.auth_token = None

    if st.session_state.authenticated:
        return True

    if not HAS_DB:
        st.warning("Login requer banco de dados. Execute o sync e ative o banco em Config.")
        return True

    init_db()
    if not tem_usuarios():
        st.markdown("### Criar primeiro usuário (administrador)")
        st.caption("Nenhum usuário existe. Crie o administrador inicial.")
        with st.form("primeiro_usuario"):
            u = st.text_input("Usuário", key="fu_user")
            p = st.text_input("Senha", type="password", key="fu_pass")
            if st.form_submit_button("Criar"):
                if u.strip() and len(p) >= 4:
                    if criar_usuario(u.strip(), p, is_admin=True):
                        user = {"username": u.strip().lower(), "is_admin": True}
                        st.session_state.authenticated = True
                        st.session_state.user = user
                        _salvar_sessao_cookie(user)
                        st.success("Usuário criado! Redirecionando...")
                        st.rerun()
                    else:
                        st.error("Erro ao criar. Usuário já existe?")
                else:
                    st.error("Usuário e senha (mín. 4 caracteres) obrigatórios.")
        return False

    # Verificar sessão salva no cookie (login persistente)
    if _restaurar_sessao_cookie():
        st.rerun()

    st.markdown("### Login")
    with st.form("login"):
        u = st.text_input("Usuário", key="login_user")
        p = st.text_input("Senha", type="password", key="login_pass")
        if st.form_submit_button("Entrar"):
            user = verificar_login(u, p)
            if user:
                st.session_state.authenticated = True
                st.session_state.user = user
                _salvar_sessao_cookie(user)
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")
    return False


def _audit(
    acao: str,
    entidade: str,
    chave: str = "",
    chamado_id: str = "",
    antes: dict | None = None,
    depois: dict | None = None,
    detalhes: str = "",
) -> None:
    """Wrapper seguro de auditoria (não quebra fluxo principal)."""
    if not (HAS_DB and registrar_auditoria):
        return
    try:
        user = st.session_state.get("user", {}) or {}
        chamado_ref = str(chamado_id or st.session_state.get("chamado_id") or "").strip()
        registrar_auditoria(
            acao=acao,
            entidade=entidade,
            chave_registro=chave,
            chamado_id=chamado_ref,
            antes=antes,
            depois=depois,
            detalhes=detalhes,
            user_id=str(user.get("id", "")),
            username=str(user.get("username", "")),
            origem="app",
        )
    except Exception:
        pass


def main() -> None:
    st.set_page_config(
        page_title="Gerenciamento de Telefones",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    _init_cookies()

    if not _render_login_or_first_user():
        return

    cwd = Path.cwd()
    default_path = str(PLANILHAS_DIR / DEFAULT_FILE) if (PLANILHAS_DIR / DEFAULT_FILE).exists() else str(cwd / DEFAULT_FILE)
    rules_path = DOC_DIR / RULES_FILE if (DOC_DIR / RULES_FILE).exists() else cwd / RULES_FILE

    st.markdown(
        """
        <style>
        #MainMenu, header, footer { visibility: hidden; }
        header[data-testid="stHeader"] { display: none; }
        div[data-testid="stToolbar"] { display: none !important; }
        section[data-testid="stSidebar"] > div:first-child { display: none; }
        .block-container { padding-top: 0 !important; padding-bottom: 1rem !important; }
        [data-testid="stHorizontalBlock"]:first-of-type {
            background: linear-gradient(90deg, #f3c300 0%, #ffcf22 100%);
            padding: 10px 20px;
            border-radius: 8px;
            margin-bottom: 1rem;
            flex-wrap: nowrap !important;
            gap: 30px !important;
            border: 1px solid #b78900;
        }
        [data-testid="stHorizontalBlock"]:first-of-type label,
        [data-testid="stHorizontalBlock"]:first-of-type p,
        [data-testid="stHorizontalBlock"]:first-of-type span,
        [data-testid="stHorizontalBlock"]:first-of-type strong,
        [data-testid="stHorizontalBlock"]:first-of-type div[data-testid="stMarkdown"],
        [data-testid="stHorizontalBlock"]:first-of-type button {
            color: #171717 !important;
            opacity: 1 !important;
        }
        [data-testid="stHorizontalBlock"]:first-of-type > div {
            overflow: visible !important;
        }
        /* Segmento em duplas (2 colunas) na navbar */
        [data-testid="stHorizontalBlock"]:first-of-type > div:nth-child(3) [role="radiogroup"] {
            display: grid !important;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            column-gap: 14px;
            row-gap: 2px;
        }
        [data-testid="stHorizontalBlock"]:first-of-type > div:nth-child(3) [role="radiogroup"] label {
            white-space: nowrap !important;
        }
        [data-testid="stHorizontalBlock"]:first-of-type > div:nth-child(3) [role="radiogroup"] p {
            white-space: nowrap !important;
            word-break: keep-all !important;
        }
        [data-testid="stHorizontalBlock"]:first-of-type button {
            background-color: transparent !important;
            border-color: rgba(23,23,23,0.75) !important;
            white-space: nowrap !important;
            min-width: fit-content !important;
        }
        .main-title { font-size: 1.8rem; font-weight: 600; color: #1f1f1f !important; margin-bottom: 0; opacity: 1 !important; }
        .main-subtitle { font-size: 0.95rem; color: #5a6c7d !important; margin-bottom: 1rem; opacity: 1 !important; }
        .team-header {
            background: linear-gradient(90deg, #2d5a24 0%, #3d7a34 100%);
            color: #ffffff !important;
            padding: 12px 16px;
            border-radius: 8px;
            margin: 24px 0 4px 0;
            font-weight: 600;
            font-size: 1.1rem;
            opacity: 1 !important;
        }
        .team-header-inactive {
            background: linear-gradient(90deg, #8b1e1e 0%, #b83434 100%);
            color: #ffffff !important;
            padding: 12px 16px;
            border-radius: 8px;
            margin: 24px 0 4px 0;
            font-weight: 600;
            font-size: 1.1rem;
            opacity: 1 !important;
        }
        .tbl-equipe {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 24px;
            font-size: 0.9rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
            border-radius: 8px;
            overflow: hidden;
        }
        .tbl-equipe thead tr { background: #191919; color: #ffd64a !important; font-weight: 600; opacity: 1 !important; }
        .tbl-equipe thead th { color: #ffffff !important; opacity: 1 !important; }
        .tbl-equipe th, .tbl-equipe td {
            padding: 10px 12px;
            text-align: left;
            border-bottom: 1px solid #e8ecf0;
            white-space: nowrap;
            opacity: 1 !important;
        }
        .tbl-equipe tbody tr:nth-child(even) { background: #f8faf8; }
        .tbl-equipe tbody tr:hover { background: #eef5ee; }
        .tbl-equipe tr.row-supervisor { background: #e8f0e8 !important; }
        .tbl-wrapper { overflow-x: auto; margin-bottom: 24px; }
        .navbar-top {
            background: linear-gradient(90deg, #1e3a5f 0%, #2d5a7b 100%);
            color: white;
            padding: 14px 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 16px;
            flex-wrap: wrap;
        }
        .navbar-top .brand { font-weight: 700; font-size: 1.15rem; white-space: nowrap; }
        .navbar-top span { white-space: nowrap; }
        div[data-testid="stSidebar"], section[data-testid="stSidebar"] { display: none; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    # Navbar no topo da página
    col_brand, col_modo, col_seg, col_cfg, col_logout = st.columns([1.6, 1.4, 2.4, 1.1, 1.3])
    query_params = st.query_params
    chamado_qp = ""
    for qp_key in ("chamado_id", "id_chamado", "ticket_id", "chamado"):
        raw_qp = query_params.get(qp_key)
        if isinstance(raw_qp, list):
            raw_qp = raw_qp[0] if raw_qp else ""
        raw_qp = str(raw_qp or "").strip()
        if raw_qp:
            chamado_qp = raw_qp
            break
    if chamado_qp:
        st.session_state["chamado_id"] = chamado_qp
    elif "chamado_id" not in st.session_state:
        st.session_state["chamado_id"] = ""
    if "_qp_hydrated" not in st.session_state:
        modo_qp = query_params.get("modo")
        seg_qp = query_params.get("segmento")
        equipe_qp = query_params.get("equipe")
        busca_qp = query_params.get("busca")
        if modo_qp in ["Linhas ativas", "Linhas desativadas"]:
            st.session_state["nav_modo"] = modo_qp
        if seg_qp in ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]:
            st.session_state["nav_segmento"] = seg_qp
        if equipe_qp:
            st.session_state["filtro_equipe"] = equipe_qp
        if busca_qp is not None:
            st.session_state["filtro_busca"] = busca_qp
        st.session_state["_qp_hydrated"] = True
    pagina_qp = query_params.get("pagina", "painel")
    current_page = pagina_qp if pagina_qp in ["painel", "config"] else "painel"
    with col_brand:
        logo_path = Path("assets/logo.png")
        if logo_path.exists():
            logo_b64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
            st.markdown(
                f"""
                <div style="display:flex; flex-direction:column; align-items:center; justify-content:center; text-align:center; margin-top:2px;">
                    <img src="data:image/png;base64,{logo_b64}" style="width:64px; height:auto; object-fit:contain; margin-bottom:4px;" />
                    <p style="font-weight:700; color:#1a1a1a; margin:0; white-space:nowrap;">Gerenciamento de Telefones</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown('<p style="font-weight:700; color:#1a1a1a; margin:0; white-space:nowrap;">Gerenciamento de Telefones</p>', unsafe_allow_html=True)
    with col_modo:
        modos_opts = ["Linhas ativas", "Linhas desativadas"]
        modo_pref = st.session_state.get("nav_modo")
        if modo_pref in modos_opts and st.session_state.get("nav_modo") != modo_pref:
            st.session_state["nav_modo"] = modo_pref
        modo_index = modos_opts.index(modo_pref) if modo_pref in modos_opts else 0
        modo_sel = st.radio("Modo", options=modos_opts, index=modo_index, horizontal=True, key="nav_modo")
    with col_seg:
        segmentos_opts = ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]
        segmento_pref = st.session_state.get("pending_segmento") or st.session_state.get("nav_segmento")
        # Se existir segmento pendente (linha recém-criada), aplica antes de instanciar o widget.
        if segmento_pref in segmentos_opts and st.session_state.get("nav_segmento") != segmento_pref:
            st.session_state["nav_segmento"] = segmento_pref
        seg_index = segmentos_opts.index(segmento_pref) if segmento_pref in segmentos_opts else 0
        segmento_sel = st.radio("Segmento", options=segmentos_opts, index=seg_index, horizontal=True, key="nav_segmento")
    modo_db = "ativas" if modo_sel == "Linhas ativas" else "desativadas"
    segmento_sem_filtro_modo = segmento_sel in ("Manutenção", "Roubo e Perda")
    st.query_params["modo"] = modo_sel
    st.query_params["segmento"] = segmento_sel
    def _render_config_content() -> None:
        nonlocal modo_sel, file_path, relacao_path
        if st.session_state.get("user", {}).get("is_admin"):
            with st.expander("Gerenciar usuários", expanded=False):
                nu = st.text_input("Novo usuário", key="new_user")
                np = st.text_input("Senha", type="password", key="new_pass")
                is_adm = st.checkbox("Administrador", key="new_admin")
                if st.button("Criar usuário", key="btn_new_user"):
                    if nu.strip() and len(np) >= 4:
                        if criar_usuario(nu.strip(), np, is_admin=is_adm):
                            st.success("Usuário criado!")
                            st.rerun()
                        else:
                            st.error("Usuário já existe.")
                    else:
                        st.error("Usuário e senha (mín. 4 caracteres) obrigatórios.")
                users = listar_usuarios()
                if users:
                    st.caption("Usuários cadastrados")
                    for usr in users:
                        adm = " (admin)" if usr["is_admin"] else ""
                        c1, c2 = st.columns([2, 1])
                        with c1:
                            st.text(f"{usr['username']}{adm}")
                        with c2:
                            if st.button("🗑️ Excluir", key=f"del_{usr['username']}", help=f"Remover {usr['username']}"):
                                excluir_usuario(usr["username"])
                                if usr["username"] == st.session_state.get("user", {}).get("username"):
                                    token = st.session_state.get("auth_token")
                                    if token:
                                        encerrar_sessao(token)
                                    if HAS_COOKIES and _cookies.ready():
                                        try:
                                            del _cookies["auth_token"]
                                            _cookies.save()
                                        except Exception:
                                            pass
                                    st.session_state.authenticated = False
                                    st.session_state.user = None
                                    st.session_state.auth_token = None
                                st.rerun()
            st.divider()
        st.markdown("**Configurações gerais**")
        use_db_cb = st.checkbox("Usar banco de dados", value=st.session_state.get("use_db", True), key="cfg_use_db", help="Dados do DB (mais rápido). Desmarque para usar planilhas.")
        if use_db_cb:
            st.session_state.use_db = True
        else:
            st.session_state.use_db = False
        if HAS_DB:
            st.caption("Banco: " + str(get_db_path()))
            if st.button("Sincronizar banco (planilhas → DB)"):
                with st.spinner("Sincronizando..."):
                    r = subprocess.run([sys.executable, "-m", "scripts.sync_db"], cwd=cwd, capture_output=True, text=True)
                if r.returncode == 0:
                    _audit(
                        acao="sincronizar_banco",
                        entidade="sync",
                        detalhes="Sync planilhas -> DB executado com sucesso",
                    )
                    st.success("Sincronização concluída!")
                    st.rerun()
                else:
                    st.error(r.stderr or r.stdout or "Erro na sincronização")
        if st.session_state.get("user", {}).get("is_admin"):
            st.divider()
            with st.expander("Histórico de alterações (auditoria)", expanded=False):
                audit_limit = st.number_input(
                    "Quantidade de registros",
                    min_value=20,
                    max_value=2000,
                    value=200,
                    step=20,
                    key="audit_limit",
                )
                logs = listar_auditoria(limit=int(audit_limit)) if listar_auditoria else []
                if not logs:
                    st.caption("Sem registros de auditoria.")
                else:
                    def _parse_json_safe(value: Any) -> Any:
                        if value is None:
                            return None
                        if isinstance(value, float) and pd.isna(value):
                            return None
                        if isinstance(value, (dict, list, int, float, bool)):
                            return value
                        txt = str(value).strip()
                        if not txt or txt.lower() in {"none", "nan"}:
                            return None
                        try:
                            return json.loads(txt)
                        except Exception:
                            return txt

                    def _friendly_action_name(action: str) -> str:
                        action_map = {
                            "salvar_edicoes": "Salvar edições",
                            "criar_linha": "Criar linha",
                            "excluir_linha": "Excluir linha",
                            "mover_equipe": "Mover equipe",
                            "mover_setor": "Mover setor",
                            "mudar_modo_linha": "Ativar/Desativar linha",
                            "enviar_manutencao": "Enviar para manutenção",
                            "sincronizar_banco": "Sincronizar banco",
                            "login": "Login",
                            "logout": "Logout",
                        }
                        action_norm = str(action or "").strip()
                        return action_map.get(action_norm, action_norm.replace("_", " ").title())

                    def _build_what_edited(row: pd.Series) -> str:
                        acao_ev = str(row.get("acao", "") or "").strip()
                        antes_ev = _parse_json_safe(row.get("antes_json"))
                        depois_ev = _parse_json_safe(row.get("depois_json"))
                        detalhes_ev = str(row.get("detalhes", "") or "").strip()
                        chave_ev = str(row.get("chave_registro", "") or "").strip()

                        if acao_ev == "salvar_edicoes" and isinstance(depois_ev, dict):
                            alteracoes: list[dict[str, Any]] = []
                            total_alteracoes = 0
                            if isinstance(antes_ev, dict):
                                raw_alter = antes_ev.get("alteracoes") or []
                                if isinstance(raw_alter, list):
                                    alteracoes = raw_alter
                                total_alteracoes = int(antes_ev.get("alteracoes_total") or 0)
                            if not alteracoes and isinstance(depois_ev, dict):
                                raw_alter = depois_ev.get("alteracoes") or []
                                if isinstance(raw_alter, list):
                                    alteracoes = raw_alter
                                total_alteracoes = int(depois_ev.get("alteracoes_total") or total_alteracoes or 0)

                            def _fmt_val(v: Any) -> str:
                                txt = "—" if v is None else str(v).strip()
                                if not txt:
                                    txt = "—"
                                return txt if len(txt) <= 36 else (txt[:33] + "...")

                            if alteracoes:
                                partes: list[str] = []
                                for ch in alteracoes[:2]:
                                    if not isinstance(ch, dict):
                                        continue
                                    ln = str(ch.get("linha") or "—").strip() or "—"
                                    campo = str(ch.get("campo") or "—").strip() or "—"
                                    antes_v = _fmt_val(ch.get("antes"))
                                    depois_v = _fmt_val(ch.get("depois"))
                                    partes.append(f"Linha {ln} | {campo}: {antes_v} -> {depois_v}")
                                if partes:
                                    total_base = total_alteracoes if total_alteracoes > 0 else len(alteracoes)
                                    extra = max(0, total_base - len(partes))
                                    sufixo = f" (+{extra} alteração(ões))" if extra > 0 else ""
                                    return " ; ".join(partes) + sufixo

                            segmento = str(depois_ev.get("segmento", "") or "").strip() or "—"
                            modo_ev = str(depois_ev.get("modo", "") or "").strip() or "—"
                            linhas_editadas: list[str] = []
                            campos_alterados: list[str] = []
                            if isinstance(antes_ev, dict):
                                linhas_editadas = antes_ev.get("linhas_editadas") or []
                                campos_alterados = antes_ev.get("campos_alterados") or []
                            if not campos_alterados and isinstance(depois_ev, dict):
                                campos_alterados = depois_ev.get("campos_alterados") or []
                            qtd_linhas = len(linhas_editadas) if isinstance(linhas_editadas, list) else 0
                            campos_validos = [str(c).strip() for c in campos_alterados if str(c).strip()]
                            campos_preview = ", ".join(campos_validos[:3])
                            if len(campos_validos) > 3:
                                campos_preview += f" +{len(campos_validos) - 3}"
                            if qtd_linhas == 1 and isinstance(linhas_editadas, list) and linhas_editadas:
                                alvo = f"linha {str(linhas_editadas[0]).strip()}"
                            else:
                                alvo = f"{qtd_linhas} linha(s)"
                            if campos_preview:
                                return f"Alterou {alvo} em {segmento} ({modo_ev}) | Campos: {campos_preview}"
                            return f"Alterou {alvo} em {segmento} ({modo_ev})"
                        if acao_ev == "criar_linha":
                            return f"Criou a linha {chave_ev or '—'}"
                        if acao_ev == "excluir_linha":
                            return f"Excluiu a linha {chave_ev or '—'}"
                        if acao_ev == "mover_equipe":
                            eq_dest = ""
                            if isinstance(depois_ev, dict):
                                eq_dest = str(depois_ev.get('EquipePadrao') or depois_ev.get('Equipe') or "").strip()
                            return f"Moveu a linha {chave_ev or '—'} para equipe {eq_dest or '—'}"
                        if acao_ev == "mover_setor":
                            seg_dest = ""
                            if isinstance(depois_ev, dict):
                                seg_dest = str(depois_ev.get('Segmento') or "").strip()
                            return f"Moveu a linha {chave_ev or '—'} para o setor {seg_dest or '—'}"
                        if acao_ev == "mudar_modo_linha":
                            modo_dest = ""
                            if isinstance(depois_ev, dict):
                                modo_dest = str(depois_ev.get('modo') or "").strip()
                            return f"Alterou a linha {chave_ev or '—'} para {modo_dest or '—'}"
                        if acao_ev == "enviar_manutencao":
                            return f"Enviou para manutenção: {chave_ev or '—'}"
                        if acao_ev in ("sincronizar_banco", "login", "logout"):
                            return f"{_friendly_action_name(acao_ev)}"

                        if detalhes_ev:
                            return detalhes_ev
                        return _friendly_action_name(acao_ev) or "Edição"

                    def _extract_chamado_id(row: pd.Series) -> str:
                        raw = row.get("chamado_id", "")
                        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                            return "—"
                        txt = str(raw).strip()
                        if not txt or txt.lower() == "nan":
                            return "—"
                        return txt

                    def _to_local_datetime(series: pd.Series) -> pd.Series:
                        # `criado_em` no SQLite vem em UTC; converter para horário local melhora a leitura.
                        dt_utc = pd.to_datetime(series, errors="coerce", utc=True)
                        tz_name = os.environ.get("APP_TIMEZONE", "America/Sao_Paulo")
                        try:
                            return dt_utc.dt.tz_convert(tz_name)
                        except Exception:
                            return dt_utc

                    df_logs = pd.DataFrame(logs)
                    acoes_edicao = {
                        "salvar_edicoes",
                        "criar_linha",
                        "excluir_linha",
                        "mover_equipe",
                        "mover_setor",
                        "mudar_modo_linha",
                        "enviar_manutencao",
                    }
                    if "acao" in df_logs.columns:
                        df_logs = df_logs[df_logs["acao"].isin(acoes_edicao)]
                    if "criado_em" in df_logs.columns:
                        df_logs["__ord_dt"] = _to_local_datetime(df_logs["criado_em"])
                        sort_cols = ["__ord_dt"] + (["id"] if "id" in df_logs.columns else [])
                        sort_asc = [False] * len(sort_cols)
                        df_logs = df_logs.sort_values(sort_cols, ascending=sort_asc, kind="stable")
                    elif "id" in df_logs.columns:
                        df_logs = df_logs.sort_values(["id"], ascending=[False], kind="stable")
                    df_logs = df_logs.reset_index(drop=True)
                    if df_logs.empty:
                        st.caption("Sem alterações de edição registradas.")
                    else:
                        df_view = pd.DataFrame()
                        if "username" in df_logs.columns:
                            df_view["Quem editou"] = df_logs["username"].fillna("").astype(str).str.strip().replace("", "—")
                        else:
                            df_view["Quem editou"] = "—"
                        df_view["Chamado"] = df_logs.apply(_extract_chamado_id, axis=1)
                        df_view["O que editou"] = df_logs.apply(_build_what_edited, axis=1)
                        dt_fmt = df_logs["__ord_dt"] if "__ord_dt" in df_logs.columns else _to_local_datetime(df_logs.get("criado_em"))
                        df_view["Quando editou"] = dt_fmt.dt.strftime("%d/%m/%Y %H:%M:%S").fillna("—")
                        st.dataframe(df_view[["Quem editou", "Chamado", "O que editou", "Quando editou"]], width="stretch", hide_index=True)
        st.divider()
        st.caption("Planilhas (quando banco não usado)")
        if modo_sel == "Linhas ativas":
            default_path = str(PLANILHAS_DIR / DEFAULT_FILE) if (PLANILHAS_DIR / DEFAULT_FILE).exists() else str(cwd / DEFAULT_FILE)
        else:
            default_path = str(PLANILHAS_DIR / DEFAULT_FILE_COMPLETO) if (PLANILHAS_DIR / DEFAULT_FILE_COMPLETO).exists() else str(cwd / DEFAULT_FILE_COMPLETO)
        planilha_key = f"cfg_planilha_{modo_sel.replace(' ', '_')}"
        if planilha_key not in st.session_state:
            st.session_state[planilha_key] = default_path
        file_path = st.text_input(
            "Planilha",
            value=st.session_state[planilha_key],
            help="Linhas ativas: SomenteAtivas. Linhas desativadas: Telefones11.25.xlsx completo.",
            key=planilha_key,
        )
        relacao_default = str(PLANILHAS_DIR / RELACAO_FILE) if (PLANILHAS_DIR / RELACAO_FILE).exists() else str(cwd / RELACAO_FILE)
        if "cfg_relacao" not in st.session_state:
            st.session_state["cfg_relacao"] = relacao_default
        relacao_path = st.session_state["cfg_relacao"]
        if modo_sel == "Linhas desativadas":
            relacao_path = st.text_input("Relação (ativas)", value=st.session_state["cfg_relacao"], help="Planilha com linhas ativas", key="cfg_relacao")

    if modo_sel == "Linhas ativas":
        file_path = st.session_state.get("cfg_planilha_Linhas_ativas")
        if not file_path:
            file_path = str(PLANILHAS_DIR / DEFAULT_FILE) if (PLANILHAS_DIR / DEFAULT_FILE).exists() else str(cwd / DEFAULT_FILE)
    else:
        file_path = st.session_state.get("cfg_planilha_Linhas_desativadas")
        if not file_path:
            file_path = str(PLANILHAS_DIR / DEFAULT_FILE_COMPLETO) if (PLANILHAS_DIR / DEFAULT_FILE_COMPLETO).exists() else str(cwd / DEFAULT_FILE_COMPLETO)
    relacao_path = st.session_state.get("cfg_relacao")
    if not relacao_path:
        relacao_path = str(PLANILHAS_DIR / RELACAO_FILE) if (PLANILHAS_DIR / RELACAO_FILE).exists() else str(cwd / RELACAO_FILE)

    with col_cfg:
        if "use_db" not in st.session_state:
            st.session_state.use_db = True
        use_db = HAS_DB and st.session_state.get("use_db", True)
        if current_page == "config":
            if st.button("⬅ Painel", key="btn_go_panel"):
                st.query_params["pagina"] = "painel"
                st.rerun()
        else:
            with st.popover("➕ Adicionar"):
                if use_db:
                    st.caption("Passo 1: escolha Segmento e Equipe. Passo 2: complete os dados na tabela.")
                    seg_add = st.selectbox("Segmento", ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"], key="add_seg")
                    eq_ref = (
                        EQUIPES_ALIMENTO
                        if seg_add == "Alimento"
                        else (
                            EQUIPES_MEDICAMENTO
                            if seg_add == "Medicamento"
                            else (
                                EQUIPES_PROMOTORES
                                if seg_add == "Promotores"
                                else (
                                    EQUIPES_INTERNOS
                                    if seg_add == "Internos"
                                    else (EQUIPES_MANUTENCAO if seg_add == "Manutenção" else EQUIPES_ROUBO_PERDA)
                                )
                            )
                        )
                    )
                    eq_add = st.selectbox("Equipe", eq_ref, key="add_equipe")

                    if st.button("Criar linha na equipe", key="add_criar_linha"):
                        try:
                            df_atual = load_linhas(modo=modo_db)
                            cols = df_atual.columns.tolist() or [
                                "Codigo", "Nome", "Equipe", "EquipePadrao", "GrupoEquipe", "TipoEquipe",
                                "Localidade", "Gestor", "Supervisor", "Segmento", "Papel", "Linha",
                                "E-mail", "Gerenciamento", "IMEI A", "IMEI B", "CHIP", "Aparelho", "Modelo",
                                "Setor", "Cargo", "Desconto", "Perfil", "Empresa", "Ativo",
                                "Numero de Serie", "Operadora", "Aba",
                            ]
                            linha_temp = f"NOVA-{secrets.token_hex(3).upper()}"
                            new_row = {c: "" for c in cols}
                            new_row.update({
                                "Segmento": seg_add,
                                "EquipePadrao": eq_add,
                                "Equipe": eq_add,
                                "GrupoEquipe": seg_add,
                                "TipoEquipe": "",
                                "Linha": linha_temp,
                            })
                            df_novo = pd.concat([df_atual, pd.DataFrame([new_row])], ignore_index=True)
                            df_novo = df_novo.sort_values(
                                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                                kind="stable",
                            )
                            save_linhas(df_novo, modo=modo_db)
                            _audit(
                                acao="criar_linha",
                                entidade="linhas",
                                chave=linha_temp,
                                depois={
                                    "segmento": seg_add,
                                    "equipe": eq_add,
                                    "modo": modo_db,
                                    "linha": linha_temp,
                                },
                                detalhes="Linha criada via botão Adicionar",
                            )
                            st.session_state.pending_segmento = seg_add
                            st.session_state.pending_equipe = eq_add
                            st.session_state.pending_linha = linha_temp
                            st.session_state.filtro_busca = ""
                            st.success("Linha criada! Complete os dados na tabela abaixo.")
                            st.rerun()
                        except Exception as exc:
                            st.error(f"Erro ao criar linha: {exc}")
                else:
                    st.caption("Ative *Usar banco de dados* em Config para adicionar usuários.")
            if st.button("⚙ Configuração", key="btn_go_config"):
                st.query_params["pagina"] = "config"
                st.rerun()
    with col_logout:
        user_display = st.session_state.get("user", {}).get("username", "")
        st.markdown(
            f'<span style="color: #1a1a1a; font-size: 0.9rem; white-space: nowrap;">{user_display}</span>&nbsp;&nbsp;',
            unsafe_allow_html=True,
        )
        if st.button("Sair", key="btn_logout"):
            token = st.session_state.get("auth_token")
            if token and HAS_DB:
                encerrar_sessao(token)
            if HAS_COOKIES and _cookies.ready():
                try:
                    del _cookies["auth_token"]
                    _cookies.save()
                except Exception:
                    pass
            st.session_state.authenticated = False
            st.session_state.user = None
            st.session_state.auth_token = None
            st.rerun()

    if current_page == "config":
        st.markdown("### ⚙ Configurações")
        _render_config_content()
        return

    # Carregar dados: banco ou planilhas
    use_db = HAS_DB and st.session_state.get("use_db", True)
    dados_do_banco = False
    if HAS_DB and use_db and (has_data(modo_db) or (segmento_sem_filtro_modo and (has_data("ativas") or has_data("desativadas")))):
        try:
            if segmento_sem_filtro_modo:
                df_ativas = load_linhas(modo="ativas") if has_data("ativas") else pd.DataFrame()
                df_des = load_linhas(modo="desativadas") if has_data("desativadas") else pd.DataFrame()
                df = pd.concat([df_ativas, df_des], ignore_index=True)
            else:
                df = load_linhas(modo=modo_db)
            dados_do_banco = True
        except Exception as exc:
            st.error(f"Erro ao ler o banco: {exc}")
            st.stop()
    if not dados_do_banco:
        if HAS_DB and use_db:
            st.info("Banco vazio. Execute **Sincronizar banco** em ⚙ Config ou desmarque para usar planilhas.")
            use_db = False
        if not use_db:
            if modo_sel == "Linhas desativadas" and (not segmento_sem_filtro_modo) and "SomenteAtivas" in file_path:
                file_path = str(PLANILHAS_DIR / DEFAULT_FILE_COMPLETO) if (PLANILHAS_DIR / DEFAULT_FILE_COMPLETO).exists() else str(cwd / DEFAULT_FILE_COMPLETO)
            if not Path(file_path).exists():
                st.error("Arquivo não encontrado. Clique em ⚙️ Config para ajustar.")
                st.stop()
            if modo_sel == "Linhas desativadas" and (not segmento_sem_filtro_modo) and not Path(relacao_path).exists():
                st.error("Arquivo da Relação não encontrado. Clique em ⚙️ Config para ajustar.")
                st.stop()
            try:
                if segmento_sem_filtro_modo:
                    fonte = str(PLANILHAS_DIR / DEFAULT_FILE_COMPLETO) if (PLANILHAS_DIR / DEFAULT_FILE_COMPLETO).exists() else str(cwd / DEFAULT_FILE_COMPLETO)
                    df = load_active_lines(fonte)
                elif modo_sel == "Linhas ativas":
                    df = load_active_lines(file_path)
                else:
                    df = load_inactive_lines(file_path, relacao_path)
                df = apply_team_standardization(df, rules_path)
            except Exception as exc:
                st.error(f"Erro ao ler a planilha: {exc}")
                st.stop()

    prefixo = "Linhas" if segmento_sem_filtro_modo else ("Linhas Ativas" if modo_sel == "Linhas ativas" else "Linhas Desativadas")
    df_full_mode = df.copy()
    # ID técnico por linha para merge seguro no data_editor (evita depender da coluna Linha).
    df_full_mode = df_full_mode.reset_index(drop=True)
    df_full_mode["__row_key"] = [f"{modo_db}:{i}" for i in range(len(df_full_mode))]
    df = df.reset_index(drop=True).copy()
    df["__row_key"] = df_full_mode["__row_key"].values
    if segmento_sel == "Alimento":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "alimento"]
        if (not dados_do_banco) and "Aba" in df.columns:
            aba_norm = df["Aba"].fillna("").astype(str).str.strip()
            df = df[aba_norm.str.lower().isin([a.lower() for a in ABAS_ALIMENTO])]
        equipes_ref = EQUIPES_ALIMENTO
        titulo = f"{prefixo} — Alimento"
    elif segmento_sel == "Medicamento":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "medicamento"]
        equipes_ref = EQUIPES_MEDICAMENTO
        titulo = f"{prefixo} — Medicamento"
    elif segmento_sel == "Promotores":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "promotores"]
        equipes_ref = EQUIPES_PROMOTORES
        titulo = f"{prefixo} — Promotores"
    elif segmento_sel == "Internos":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "internos"]
        equipes_ref = EQUIPES_INTERNOS
        titulo = f"{prefixo} — Internos"
    elif segmento_sel == "Manutenção":
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "manutenção"]
        equipes_ref = EQUIPES_MANUTENCAO
        titulo = f"{prefixo} — Manutenção"
    else:
        df = df[df["Segmento"].fillna("").astype(str).str.strip().str.lower() == "roubo e perda"]
        equipes_ref = EQUIPES_ROUBO_PERDA
        titulo = f"{prefixo} — Roubo e Perda"

    if df.empty:
        st.warning(f"Nenhum dado de {segmento_sel} encontrado.")
        st.stop()

    if dados_do_banco:
        tbl = df.copy()
        tbl = tbl.sort_values(
            ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
            kind="stable",
        )
    else:
        tbl = build_full_table(df)
    # Normaliza valores nulos para não exibir "None" no editor/tabelas.
    tbl = tbl.replace({None: ""}).fillna("")
    if segmento_sel == "Roubo e Perda":
        cols_front = [
            "Data Ocorrência",
            "Data Solicitação TBS",
            "Nome de Guerra",
            "Nome",
            "Linha",
            "Motivo",
            "Observação",
            "Marca",
            "Modelo",
            "IMEI A",
            "IMEI B",
            "CHIP",
            "Patrimonio",
            "Numero de Serie",
            "Aba",
        ]
        existing_front = [c for c in cols_front if c in tbl.columns]
        tail = [c for c in tbl.columns if c not in existing_front]
        tbl = tbl[existing_front + tail]
    elif segmento_sel == "Manutenção":
        cols_front = [
            "Data da Troca",
            "Data Retorno",
            "Codigo",
            "Nome",
            "Linha",
            "Aparelho",
            "Modelo",
            "IMEI A",
            "IMEI B",
            "CHIP",
            "Numero de Serie",
            "Aba",
        ]
        existing_front = [c for c in cols_front if c in tbl.columns]
        tail = [c for c in tbl.columns if c not in existing_front]
        tbl = tbl[existing_front + tail]
    if not dados_do_banco and segmento_sel == "Alimento":
        gerentes_set = _get_gerentes_alimento(rules_path)
        mask_papel = tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente"
        mask_nome = tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_set
        )
        mask_gerente = mask_papel | mask_nome
        if mask_gerente.any():
            tbl.loc[mask_gerente, "EquipePadrao"] = "Gerentes do Alimento"
            tbl = tbl.sort_values(
                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                kind="stable",
            )
    if not dados_do_banco and segmento_sel == "Medicamento":
        gerentes_set = _get_gerentes_medicamento(rules_path)
        mask_papel = tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente"
        mask_nome = tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_set
        )
        mask_gerente = mask_papel | mask_nome
        if mask_gerente.any():
            tbl.loc[mask_gerente, "EquipePadrao"] = "Gerentes do Medicamento"
            tbl = tbl.sort_values(
                ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"],
                kind="stable",
            )
    equipes_validas = sorted(set(tbl["EquipePadrao"].dropna().astype(str).str.strip().unique()))
    equipes = [e for e in equipes_ref if e in equipes_validas]
    equipes += [e for e in equipes_validas if e not in equipes]
    tbl_segment = tbl.copy()

    st.markdown(f'<p class="main-title">{titulo}</p>', unsafe_allow_html=True)
    st.markdown('<p class="main-subtitle">Controle de linhas telefônicas por equipe. Selecione uma equipe para ver os detalhes.</p>', unsafe_allow_html=True)

    col_filtro, col_busca = st.columns([1, 3])
    with col_filtro:
        equipe_options = ["Todas as equipes"] + equipes
        pending_equipe = st.session_state.get("pending_equipe")
        if pending_equipe in equipe_options:
            st.session_state.filtro_equipe = pending_equipe
        equipe_sel = st.selectbox(
            "Equipe",
            options=equipe_options,
            format_func=lambda x: "Todas as equipes" if x == "Todas as equipes" else x,
            label_visibility="collapsed",
            key="filtro_equipe",
        )
    with col_busca:
        busca = st.text_input("Buscar", placeholder="Nome, código, linha, aparelho...", label_visibility="collapsed", key="filtro_busca")
    st.query_params["equipe"] = equipe_sel
    st.query_params["busca"] = busca

    query = busca.strip().lower()
    if query:
        mask = pd.Series(False, index=tbl.index)
        for c in ["Nome", "Codigo", "Linha", "Aparelho", "Modelo", "EquipePadrao", "Supervisor", "Gestor"]:
            if c in tbl.columns:
                mask = mask | tbl[c].astype(str).str.lower().str.contains(query, na=False)
        tbl = tbl[mask]

    if equipe_sel != "Todas as equipes":
        tbl = tbl[tbl["EquipePadrao"] == equipe_sel]

    pending_linha = st.session_state.get("pending_linha")
    if pending_linha and not tbl.empty and "Linha" in tbl.columns:
        mask_pending = tbl["Linha"].fillna("").astype(str) == pending_linha
        if mask_pending.any():
            st.info(f"Nova linha criada na equipe **{equipe_sel}**. Procure por `{pending_linha}` e edite os dados.")
            st.session_state.pending_segmento = None
            st.session_state.pending_equipe = None
            st.session_state.pending_linha = None

    n_linhas = tbl["Linha"].nunique()
    n_pessoas = len(tbl)
    lbl_linhas = "linhas" if segmento_sem_filtro_modo else ("linhas desativadas" if modo_sel == "Linhas desativadas" else "linhas ativas")
    st.caption(f"**{n_pessoas}** registros | **{n_linhas}** {lbl_linhas}")

    if tbl.empty:
        st.info("Nenhum resultado. Tente outro filtro ou termo de busca.")
        st.stop()

    if segmento_sel == "Medicamento":
        sort_cols = ["Localidade", "Papel", "Nome", "Linha"]
        sort_ascending = [True, True, True, True]
    elif segmento_sel == "Manutenção":
        sort_cols = ["Data da Troca", "Data Retorno", "Nome", "Linha"]
        sort_ascending = [False, False, True, True]
    elif segmento_sel == "Roubo e Perda":
        sort_cols = ["Data Ocorrência", "Data da Troca", "Nome", "Linha"]
        sort_ascending = [False, False, True, True]
    else:
        sort_cols = ["Papel", "Nome", "Linha"]
        sort_ascending = [True, True, True]
    header_cls = "team-header-inactive" if modo_sel == "Linhas desativadas" else "team-header"
    usar_editor = HAS_DB and dados_do_banco
    if usar_editor:
        st.caption("✏️ Clique em qualquer célula para editar na linha. Depois clique em *Salvar alterações no banco*.")

    def _sort_for_display(df_in: pd.DataFrame) -> pd.DataFrame:
        """Ordena visualização, priorizando datas em ordem decrescente quando aplicável."""
        df_sort = df_in.copy()
        sort_eff: list[str] = []
        asc_eff: list[bool] = []
        date_cols = {"Data da Troca", "Data Retorno", "Data Ocorrência", "Data Solicitação TBS"}
        for col, asc in zip(sort_cols, sort_ascending):
            if col not in df_sort.columns:
                continue
            if col in date_cols:
                aux = "__ord_" + normalize_text(col).replace(" ", "_")
                df_sort[aux] = pd.to_datetime(df_sort[col], errors="coerce")
                sort_eff.append(aux)
            else:
                sort_eff.append(col)
            asc_eff.append(asc)
        if sort_eff:
            df_sort = df_sort.sort_values(sort_eff, kind="stable", ascending=asc_eff)
        aux_cols = [c for c in df_sort.columns if c.startswith("__ord_")]
        if aux_cols:
            df_sort = df_sort.drop(columns=aux_cols, errors="ignore")
        return df_sort

    def _render_team_actions(eq_nome: str, df_equipe: pd.DataFrame) -> None:
        """Ações por bloco de equipe: excluir, mover e ativar/desativar."""
        linhas_team = sorted([str(x).strip() for x in df_equipe["Linha"].dropna().astype(str).tolist() if str(x).strip()])
        if not linhas_team:
            return
        key_base = normalize_text(eq_nome).replace(" ", "_")
        with st.expander(f"Ações da equipe: {eq_nome}", expanded=False):
            c1, c2, c3 = st.columns([1.3, 1.4, 1.3])
            with c1:
                linha_alvo = st.selectbox("Linha", options=[""] + linhas_team, key=f"acao_linha_{key_base}")
            with c2:
                acao_linha = st.selectbox(
                    "Ação",
                    options=[
                        "Excluir linha",
                        "Mover para outra equipe",
                        "Mover para outro setor",
                        "Enviar aparelho para manutenção",
                        "Desativar linha",
                        "Ativar linha",
                    ],
                    key=f"acao_tipo_{key_base}",
                )
            destino_setor = ""
            destino_equipe = ""
            with c3:
                if acao_linha == "Mover para outra equipe":
                    destino_equipe = st.selectbox("Equipe destino", options=equipes, key=f"acao_eq_dest_{key_base}")
                elif acao_linha == "Mover para outro setor":
                    destino_setor = st.selectbox("Setor destino", options=["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"], key=f"acao_seg_dest_{key_base}")
                    eq_dest_ref = (
                        EQUIPES_ALIMENTO
                        if destino_setor == "Alimento"
                        else (
                            EQUIPES_MEDICAMENTO
                            if destino_setor == "Medicamento"
                            else (
                                EQUIPES_PROMOTORES
                                if destino_setor == "Promotores"
                                else (
                                    EQUIPES_INTERNOS
                                    if destino_setor == "Internos"
                                    else (EQUIPES_MANUTENCAO if destino_setor == "Manutenção" else EQUIPES_ROUBO_PERDA)
                                )
                            )
                        )
                    )
                    destino_equipe = st.selectbox("Equipe destino", options=eq_dest_ref, key=f"acao_eq_dest_setor_{key_base}")

            if st.button("Executar ação", key=f"acao_exec_{key_base}"):
                if not linha_alvo:
                    st.error("Selecione uma linha.")
                    return
                try:
                    modo_origem = modo_db
                    df_origem = load_linhas(modo=modo_origem)
                    mask_origem = df_origem["Linha"].fillna("").astype(str) == linha_alvo
                    if not mask_origem.any():
                        st.error("Linha não encontrada no modo atual.")
                        return
                    before_row = df_origem[mask_origem].iloc[0].to_dict()

                    if acao_linha == "Excluir linha":
                        df_origem = df_origem[~mask_origem].copy()
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="excluir_linha",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            detalhes=f"Exclusão na equipe {eq_nome}",
                        )
                        st.success("Linha excluída com sucesso.")
                        st.rerun()

                    elif acao_linha == "Mover para outra equipe":
                        df_origem.loc[mask_origem, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_origem, "Equipe"] = destino_equipe
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_equipe",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={
                                "EquipePadrao": destino_equipe,
                                "Equipe": destino_equipe,
                                "modo": modo_origem,
                            },
                            detalhes=f"Origem equipe: {eq_nome}",
                        )
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"Linha movida para equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Mover para outro setor":
                        df_origem.loc[mask_origem, "Segmento"] = destino_setor
                        df_origem.loc[mask_origem, "GrupoEquipe"] = destino_setor
                        df_origem.loc[mask_origem, "EquipePadrao"] = destino_equipe
                        df_origem.loc[mask_origem, "Equipe"] = destino_equipe
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="mover_setor",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={
                                "Segmento": destino_setor,
                                "GrupoEquipe": destino_setor,
                                "EquipePadrao": destino_equipe,
                                "Equipe": destino_equipe,
                                "modo": modo_origem,
                            },
                        )
                        st.session_state.pending_segmento = destino_setor
                        st.session_state.pending_equipe = destino_equipe
                        st.success(f"Linha movida para setor {destino_setor} / equipe {destino_equipe}.")
                        st.rerun()

                    elif acao_linha == "Enviar aparelho para manutenção":
                        linha_origem = df_origem[mask_origem].iloc[0].copy()
                        linha_manut = linha_origem.copy()
                        linha_manut["Segmento"] = "Manutenção"
                        linha_manut["GrupoEquipe"] = "Manutenção"
                        linha_manut["EquipePadrao"] = "Manutenção"
                        linha_manut["Equipe"] = "Manutenção"
                        linha_manut["Linha"] = f"MANUT-{secrets.token_hex(3).upper()}"
                        linha_manut["Nome"] = f"{str(linha_origem.get('Nome', '')).strip()} (Manutenção)".strip()
                        base_ger = str(linha_manut.get("Gerenciamento", "") or "").strip()
                        linha_manut["Gerenciamento"] = f"{base_ger} | Origem: {str(linha_origem.get('Linha', '')).strip()}".strip(" |")
                        df_origem = pd.concat([df_origem, pd.DataFrame([linha_manut])], ignore_index=True)
                        save_linhas(df_origem, modo=modo_origem)
                        _audit(
                            acao="enviar_manutencao",
                            entidade="linhas",
                            chave=str(linha_manut["Linha"]),
                            antes=before_row,
                            depois=linha_manut.to_dict(),
                            detalhes="Criada cópia para manutenção",
                        )
                        st.session_state.pending_segmento = "Manutenção"
                        st.session_state.pending_equipe = "Manutenção"
                        st.session_state.pending_linha = linha_manut["Linha"]
                        st.success("Aparelho enviado para Manutenção. Atualize a linha original com o novo aparelho.")
                        st.rerun()

                    elif acao_linha in ("Desativar linha", "Ativar linha"):
                        destino_modo = "desativadas" if acao_linha == "Desativar linha" else "ativas"
                        if destino_modo == modo_origem:
                            st.info("A linha já está neste modo.")
                            return
                        df_dest = load_linhas(modo=destino_modo)
                        movidas = df_origem[mask_origem].copy()
                        df_origem = df_origem[~mask_origem].copy()
                        linhas_movidas = set(movidas["Linha"].fillna("").astype(str))
                        if not df_dest.empty:
                            mask_rm = ~df_dest["Linha"].fillna("").astype(str).isin(linhas_movidas)
                            df_dest = df_dest[mask_rm].copy()
                        df_dest = pd.concat([df_dest, movidas], ignore_index=True)
                        save_linhas(df_origem, modo=modo_origem)
                        save_linhas(df_dest, modo=destino_modo)
                        _audit(
                            acao="mudar_modo_linha",
                            entidade="linhas",
                            chave=linha_alvo,
                            antes=before_row,
                            depois={**before_row, "modo": destino_modo},
                        )
                        st.success(f"Linha movida para {destino_modo}.")
                        st.rerun()
                except Exception as exc:
                    st.error(f"Erro ao executar ação: {exc}")

    all_edited: list[pd.DataFrame] = []

    if equipe_sel == "Todas as equipes":
        eq_order = [e for e in equipes if e in tbl["EquipePadrao"].unique()]
        eq_order += [e for e in sorted(tbl["EquipePadrao"].unique()) if e not in eq_order]
        for eq in eq_order:
            dfe = _sort_for_display(tbl[tbl["EquipePadrao"] == eq])
            if segmento_sel in ("Internos", "Manutenção", "Roubo e Perda"):
                header_text = f'<div class="{header_cls}">📋 {eq}</div>'
            elif eq in ("Gerentes do Alimento", "Gerentes do Medicamento"):
                gestor = "—"
                supervisor = "—"
                header_text = f'<div class="{header_cls}">📋 {eq} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
            else:
                gestor = non_empty_or_default(dfe["Gestor"].iloc[0], "—")
                supervisor = _supervisor_display(dfe, eq, rules_path)
                header_text = f'<div class="{header_cls}">📋 {eq} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
            st.markdown(header_text, unsafe_allow_html=True)
            if usar_editor:
                ed = st.data_editor(
                    dfe,
                    width="stretch",
                    num_rows="fixed",
                    hide_index=True,
                    key=f"ed_{eq}",
                    column_config={"__row_key": None},
                    disabled=["__row_key"],
                )
                all_edited.append(ed)
                _render_team_actions(eq, dfe)
            else:
                st.markdown(_render_equipe_tabela(dfe, eq), unsafe_allow_html=True)
    else:
        dfe = _sort_for_display(tbl)
        if segmento_sel in ("Internos", "Manutenção", "Roubo e Perda"):
            header_text = f'<div class="{header_cls}">📋 {equipe_sel}</div>'
        elif equipe_sel in ("Gerentes do Alimento", "Gerentes do Medicamento"):
            gestor = "—"
            supervisor = "—"
            header_text = f'<div class="{header_cls}">📋 {equipe_sel} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
        else:
            gestor = non_empty_or_default(dfe["Gestor"].iloc[0], "—")
            supervisor = _supervisor_display(dfe, equipe_sel, rules_path)
            header_text = f'<div class="{header_cls}">📋 {equipe_sel} — <strong>Gerente:</strong> {gestor}  |  <strong>Supervisor:</strong> {supervisor}</div>'
        st.markdown(header_text, unsafe_allow_html=True)
        if usar_editor:
            ed = st.data_editor(
                dfe,
                width="stretch",
                num_rows="fixed",
                hide_index=True,
                key=f"ed_{equipe_sel}",
                column_config={"__row_key": None},
                disabled=["__row_key"],
            )
            all_edited.append(ed)
            _render_team_actions(equipe_sel, dfe)
        else:
            st.markdown(_render_equipe_tabela(dfe, equipe_sel), unsafe_allow_html=True)

    if usar_editor and all_edited:
        if st.button("Salvar alterações no banco"):
            seg_lower = segmento_sel.strip().lower()
            mask_seg = df_full_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower
            outros = df_full_mode[~mask_seg].copy()
            segment_atual = df_full_mode[mask_seg].copy()
            edited_df = pd.concat(all_edited, ignore_index=True)
            row_keys_editadas = set(edited_df["__row_key"].dropna().astype(str).unique()) if "__row_key" in edited_df.columns else set()

            def _norm_cmp(v: Any) -> str:
                if v is None:
                    return ""
                if isinstance(v, float) and pd.isna(v):
                    return ""
                return str(v).strip()

            row_keys_alteradas: set[str] = set()
            campos_alterados_set: set[str] = set()
            linhas_editadas_lista: list[str] = []
            celulas_alteradas = 0
            alteracoes_detalhadas: list[dict[str, str]] = []
            max_alteracoes_detalhadas = 200

            if row_keys_editadas and "__row_key" in segment_atual.columns and "__row_key" in edited_df.columns:
                base_idx = segment_atual.set_index("__row_key", drop=False)
                edit_idx = edited_df.set_index("__row_key", drop=False)
                cols_compare = [c for c in edited_df.columns if c in segment_atual.columns and c != "__row_key"]
                for rk in row_keys_editadas:
                    if rk not in base_idx.index or rk not in edit_idx.index:
                        continue
                    base_row = base_idx.loc[rk]
                    edit_row = edit_idx.loc[rk]
                    linha_ref = _norm_cmp(edit_row.get("Linha")) or _norm_cmp(base_row.get("Linha")) or str(rk)
                    houve_mudanca = False
                    for col in cols_compare:
                        base_norm = _norm_cmp(base_row.get(col))
                        edit_norm = _norm_cmp(edit_row.get(col))
                        if base_norm != edit_norm:
                            houve_mudanca = True
                            celulas_alteradas += 1
                            campos_alterados_set.add(str(col))
                            if len(alteracoes_detalhadas) < max_alteracoes_detalhadas:
                                alteracoes_detalhadas.append(
                                    {
                                        "linha": linha_ref,
                                        "campo": str(col),
                                        "antes": base_norm or "—",
                                        "depois": edit_norm or "—",
                                    }
                                )
                    if houve_mudanca:
                        row_keys_alteradas.add(str(rk))
                        linhas_editadas_lista.append(linha_ref)

            linhas_editadas_lista = sorted(set(linhas_editadas_lista))
            campos_alterados_lista = sorted(campos_alterados_set)

            if not row_keys_alteradas:
                st.info("Nenhuma alteração detectada para salvar.")
            else:
                mask_rows_alteradas = edited_df["__row_key"].fillna("").astype(str).isin(row_keys_alteradas)
                edited_changed_df = edited_df[mask_rows_alteradas].copy()
                mask_manter = ~segment_atual["__row_key"].fillna("").astype(str).isin(row_keys_alteradas)
                segment_nao_editado = segment_atual[mask_manter].copy()
                novos_seg = pd.concat([segment_nao_editado, edited_changed_df], ignore_index=True)
                novos = pd.concat([outros, novos_seg], ignore_index=True)
                novos = novos.drop(columns=["__row_key"], errors="ignore")
                sort_cols_save = [c for c in ["GrupoEquipe", "EquipePadrao", "Supervisor", "Papel", "Nome", "Linha"] if c in novos.columns]
                if sort_cols_save:
                    novos = novos.sort_values(sort_cols_save, kind="stable")
                n = save_linhas(novos, modo=modo_db)
                campos_txt = ", ".join(campos_alterados_lista[:4]) if campos_alterados_lista else "campos não identificados"
                if campos_alterados_lista and len(campos_alterados_lista) > 4:
                    campos_txt += f" (+{len(campos_alterados_lista) - 4})"
                _audit(
                    acao="salvar_edicoes",
                    entidade="linhas",
                    chave=modo_db,
                    antes={
                        "registros_segmento_antes": int(len(segment_atual)),
                        "linhas_editadas": linhas_editadas_lista,
                        "campos_alterados": campos_alterados_lista,
                        "celulas_alteradas": int(celulas_alteradas),
                        "alteracoes_total": int(celulas_alteradas),
                        "alteracoes": alteracoes_detalhadas,
                    },
                    depois={
                        "registros_segmento_depois": int(len(novos_seg)),
                        "registros_gravados_modo": int(n),
                        "segmento": segmento_sel,
                        "modo": modo_db,
                        "campos_alterados": campos_alterados_lista,
                        "alteracoes_total": int(celulas_alteradas),
                        "alteracoes": alteracoes_detalhadas,
                    },
                    detalhes=f"Edição de {len(linhas_editadas_lista)} linha(s). Campos: {campos_txt}",
                )
                st.success(f"Alterações salvas! {len(linhas_editadas_lista)} linha(s) alterada(s).")
                st.rerun()

    st.markdown("---")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for segmento in ["Alimento", "Medicamento", "Promotores", "Internos", "Manutenção", "Roubo e Perda"]:
            seg_lower = segmento.strip().lower()
            mask = df_full_mode["Segmento"].fillna("").astype(str).str.strip().str.lower() == seg_lower
            df_seg = df_full_mode[mask]
            if not df_seg.empty:
                df_seg.to_excel(writer, sheet_name=segmento, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    header_fill = PatternFill(start_color="2D5A24", end_color="2D5A24", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    gerente_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    supervisor_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for ws in wb.worksheets:
        header_map = {str(cell.value or "").strip(): col_idx for col_idx, cell in enumerate(ws[1], 1)}
        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
            row_num = row[0].row
            papel_col = header_map.get("Papel")
            equipe_col = header_map.get("EquipePadrao")
            is_gerente = False
            is_supervisor = False
            if papel_col:
                papel_val = str(ws.cell(row=row_num, column=papel_col).value or "").strip().lower()
                if papel_val == "gerente":
                    is_gerente = True
                elif papel_val == "supervisor":
                    is_supervisor = True
            if equipe_col:
                eq_val = str(ws.cell(row=row_num, column=equipe_col).value or "").strip()
                if eq_val in ("Gerentes do Alimento", "Gerentes do Medicamento"):
                    is_gerente = True
            row_fill = gerente_fill if is_gerente else (supervisor_fill if is_supervisor else None)
            if row_fill:
                for c in row:
                    c.fill = row_fill
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 14
            for cell in ws[col_letter]:
                try:
                    max_len = max(max_len, min(len(str(cell.value or "")), 50))
                except TypeError:
                    pass
            ws.column_dimensions[col_letter].width = max_len
        ws.freeze_panes = "A2"
    buf_out = io.BytesIO()
    wb.save(buf_out)
    buf_out.seek(0)
    st.download_button(
        "Exportar tudo para Excel",
        data=buf_out,
        file_name="linhas_telefones.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
