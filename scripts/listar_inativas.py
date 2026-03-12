"""Lista linhas que estão em Telefones11.25.xlsx mas NÃO em Relação Linhas Prosper270226.xlsx (ListaAtual).
   Essas são as linhas que provavelmente NÃO estão mais ativas.
   Considera apenas as abas: Prosper Norte, Prosper Sul, Nova Prosper, Promotores."""
import sys
from pathlib import Path
from openpyxl import load_workbook


ABAS_FOCO = ["Prosper Norte", "Prosper Sul", "Nova Prosper", "Promotores"]


def digits_only(v):
    return "".join(c for c in str(v) if c.isdigit())


def find_line_col(ws):
    for row in range(1, min(15, ws.max_row + 1)):
        for col in range(1, min(25, ws.max_column + 1)):
            if "linha" in str(ws.cell(row, col).value or "").lower():
                return row, col
    return None, None


def main():
    root = Path(__file__).resolve().parent.parent
    planilhas = root / "data" / "planilhas"
    rel_path = planilhas / "Relação Linhas Prosper270226.xlsx" if (planilhas / "Relação Linhas Prosper270226.xlsx").exists() else root / "Relação Linhas Prosper270226.xlsx"
    tel_path = planilhas / "Telefones11.25.xlsx" if (planilhas / "Telefones11.25.xlsx").exists() else root / "Telefones11.25.xlsx"

    if not rel_path.exists():
        print("Arquivo não encontrado: Relação Linhas Prosper270226.xlsx")
        sys.exit(1)
    if not tel_path.exists():
        print("Arquivo não encontrado: Telefones11.25.xlsx")
        sys.exit(1)

    # 1. Linhas ATIVAS (Relação - ListaAtual)
    wb_rel = load_workbook(rel_path, data_only=True)
    if "ListaAtual" not in wb_rel.sheetnames:
        print("Aba 'ListaAtual' não encontrada na Relação.")
        sys.exit(1)
    ws_rel = wb_rel["ListaAtual"]
    ativas = set()
    for r in range(2, ws_rel.max_row + 1):
        v = ws_rel.cell(r, 4).value
        d = digits_only(v)
        if 10 <= len(d) <= 13:
            ativas.add(d)
    print(f"Linhas ativas (Relação ListaAtual): {len(ativas)}\n")

    # 2. Linhas em Telefones11.25 — apenas abas: Prosper Norte, Prosper Sul, Nova Prosper, Promotores
    wb_tel = load_workbook(tel_path, data_only=True)
    abas_foco_norm = {a.strip().lower() for a in ABAS_FOCO}
    todas_info = []
    for sh in wb_tel.sheetnames:
        if sh.strip().lower() not in abas_foco_norm:
            continue
        ws = wb_tel[sh]
        hr, lc = find_line_col(ws)
        if hr is None:
            continue
        for r in range(hr + 1, ws.max_row + 1):
            v = ws.cell(r, lc).value
            d = digits_only(v)
            if 10 <= len(d) <= 13:
                cod = str(ws.cell(r, 1).value or "").strip()
                nome = str(ws.cell(r, 2).value or "").strip()
                eq = str(ws.cell(r, 3).value or "").strip()
                todas_info.append(
                    {"linha": d, "codigo": cod, "nome": nome, "equipe": eq, "aba": sh}
                )
    print(f"Total linhas em Telefones11.25.xlsx (abas Prosper Norte, Sul, Nova Prosper, Promotores): {len(todas_info)}\n")

    # 3. INATIVAS (em Telefones mas NÃO na Relação)
    inativas_info = [x for x in todas_info if x["linha"] not in ativas]
    print(f"LINHAS NÃO ATIVAS (em Telefones11.25 mas não na Relação): {len(inativas_info)}\n")
    print("=" * 90)
    print(f"{'LINHA':<15} {'CÓDIGO':<12} {'NOME':<42} {'EQUIPE':<25} {'ABA'}")
    print("=" * 90)
    for x in sorted(inativas_info, key=lambda t: (t["aba"], t["nome"].upper(), t["linha"])):
        nome = (x["nome"] or "")[:40]
        eq = (x["equipe"] or "")[:24]
        print(f"{x['linha']:<15} {x['codigo']:<12} {nome:<42} {eq:<25} {x['aba']}")
    print("=" * 90)


if __name__ == "__main__":
    main()
