#!/usr/bin/env python
"""
Script para sincronizar planilhas Excel com o banco de dados.
Execute antes de usar o app para popular o banco.

Uso:
    python -m scripts.sync_db
    python scripts/sync_db.py
"""

import sys
from pathlib import Path

# Adiciona raiz do projeto ao path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.core.config import (
    DATA_DIR, PLANILHAS_DIR, DOC_DIR, get_db_path,
    DEFAULT_FILE, DEFAULT_FILE_COMPLETO, RELACAO_FILE, RULES_FILE,
    ABAS_FOCO,
)
from src.db.repository import init_db, save_linhas, save_relacao_ativas, get_connection
from src.utils.text import normalize_text, normalize_team_key, title_case_safe, digits_only
from src.utils.validators import is_valid_phone

import pandas as pd
from openpyxl import load_workbook


def _find_header_row(ws):
    from src.utils.text import normalize_text
    max_r = min(80, ws.max_row)
    max_c = min(80, ws.max_column)
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            if normalize_text(ws.cell(row, col).value) == "linha":
                return row, col
    return None, None


def _collect_headers(ws, header_row: int, start_col: int, limit: int = 40):
    headers = []
    empty_streak = 0
    max_c = min(ws.max_column, start_col + limit)
    for col in range(1, max_c + 1):
        value = ws.cell(header_row, col).value
        text = str(value).strip() if value is not None else ""
        if text == "":
            headers.append(f"col_{col}")
            empty_streak += 1
        else:
            headers.append(text)
            empty_streak = 0
        if col >= start_col and empty_streak >= 8:
            break
    return headers


def _map_known_column(column_name: str) -> str:
    c = normalize_text(column_name)
    mapping = {
        "codigo": "Codigo", "codigos": "Codigo", "asset": "Codigo", "nome": "Nome", "nomes": "Nome",
        "nome colaborador": "Nome", "nome de guerra": "Nome de Guerra",
        "equipe": "Equipe", "local": "Localidade", "localidade": "Localidade",
        "data da troca": "Data da Troca", "data troca": "Data da Troca",
        "data ocorrencia": "Data Ocorrência", "data ocorrência": "Data Ocorrência",
        "data retorno": "Data Retorno", "data de retorno": "Data Retorno",
        "data solicitacao tbs": "Data Solicitação TBS", "data solicitação tbs": "Data Solicitação TBS",
        "linha": "Linha", "email": "E-mail", "e-mail": "E-mail",
        "gerenciamento": "Gerenciamento", "bloqueio": "Gerenciamento",
        "imei": "IMEI A",
        "imei a": "IMEI A", "imei b": "IMEI B", "imei2": "IMEI B",
        "chip": "CHIP", "marca": "Marca", "aparelho": "Aparelho", "modelo": "Modelo", "setor": "Setor",
        "cargo": "Cargo", "desconto": "Desconto", "perfil": "Perfil", "empresa": "Empresa",
        "ativo": "Ativo", "ativos": "Ativo", "numero de serie": "Numero de Serie",
        "operadora": "Operadora", "s/n": "Numero de Serie", "ns": "Numero de Serie", "patrimonio": "Patrimonio",
        "motivo": "Motivo", "obs": "Observação",
    }
    return mapping.get(c, column_name)


def load_excel(file_path: Path) -> pd.DataFrame:
    """Carrega dados do Excel."""
    from src.core.config import DEFAULT_COLUMNS
    wb = load_workbook(file_path, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, line_col = _find_header_row(ws)
        if header_row is None or line_col is None:
            headers = DEFAULT_COLUMNS.copy()
            first_data_row, max_col = 1, len(headers)
        else:
            headers = [_map_known_column(h) for h in _collect_headers(ws, header_row, line_col)]
            first_data_row, max_col = header_row + 1, len(headers)
        for row_idx in range(first_data_row, ws.max_row + 1):
            line_val = ws.cell(row_idx, 4).value if header_row is None else ws.cell(row_idx, line_col).value
            if not is_valid_phone(line_val):
                continue
            row_data = {}
            seen_keys = {}
            for col_idx in range(1, max_col + 1):
                key = headers[col_idx - 1] if col_idx - 1 < len(headers) else f"col_{col_idx}"
                seen_keys[key] = seen_keys.get(key, 0) + 1
                if key == "IMEI A" and seen_keys[key] == 2:
                    key = "IMEI B"
                row_data[key] = ws.cell(row_idx, col_idx).value
            row_data["Linha"] = digits_only(line_val)
            row_data["Aba"] = sheet_name
            rows.append(row_data)
    if not rows:
        return pd.DataFrame(columns=DEFAULT_COLUMNS + ["Aba"])
    df = pd.DataFrame(rows)
    for col in DEFAULT_COLUMNS + ["Aba"]:
        if col not in df.columns:
            df[col] = ""
    df["Equipe"] = df["Equipe"].fillna("Sem Equipe").astype(str).str.strip()
    df.loc[df["Equipe"] == "", "Equipe"] = "Sem Equipe"
    df["Nome"] = df["Nome"].fillna("").astype(str).str.strip()
    df["Codigo"] = df["Codigo"].fillna("").astype(str).str.strip()
    return df


def load_relacao(file_path: Path) -> frozenset:
    """Carrega linhas ativas da Relação."""
    try:
        wb = load_workbook(file_path, data_only=True)
        if "ListaAtual" not in wb.sheetnames:
            return frozenset()
        ws = wb["ListaAtual"]
        return frozenset(
            digits_only(ws.cell(r, 4).value)
            for r in range(2, ws.max_row + 1)
            if 10 <= len(digits_only(ws.cell(r, 4).value or "")) <= 13
        )
    except Exception:
        return frozenset()


def main():
    print("Criando diretórios...")
    PLANILHAS_DIR.mkdir(parents=True, exist_ok=True)
    DOC_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Buscar arquivos: data/planilhas/ e doc/ primeiro, depois raiz (compatibilidade)
    raw_ativas = PLANILHAS_DIR / DEFAULT_FILE if (PLANILHAS_DIR / DEFAULT_FILE).exists() else ROOT / DEFAULT_FILE
    raw_completo = PLANILHAS_DIR / DEFAULT_FILE_COMPLETO if (PLANILHAS_DIR / DEFAULT_FILE_COMPLETO).exists() else ROOT / DEFAULT_FILE_COMPLETO
    relacao_path = PLANILHAS_DIR / RELACAO_FILE if (PLANILHAS_DIR / RELACAO_FILE).exists() else ROOT / RELACAO_FILE
    rules_path = DOC_DIR / RULES_FILE if (DOC_DIR / RULES_FILE).exists() else ROOT / RULES_FILE

    if not raw_ativas.exists():
        print(f"ERRO: Planilha não encontrada: {raw_ativas}")
        print("Copie os arquivos para data/planilhas/ ou para a raiz do projeto.")
        sys.exit(1)

    print("Inicializando banco...")
    init_db()

    # Importar processamento do app legado
    sys.path.insert(0, str(ROOT))
    import app as legacy_app

    print("Carregando linhas ativas...")
    df_ativas = load_excel(raw_ativas)
    ativas_set = load_relacao(relacao_path)
    mask_aba = df_ativas["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_FOCO]
    )
    mask_ativa = df_ativas["Linha"].isin(ativas_set)
    df_ativas = df_ativas[mask_aba & mask_ativa].copy()

    print("Carregando linhas desativadas...")
    df_full = load_excel(raw_completo)
    mask_aba_full = df_full["Aba"].fillna("").astype(str).str.strip().str.lower().isin(
        [a.lower() for a in ABAS_FOCO]
    )
    mask_inativa = ~df_full["Linha"].isin(ativas_set)
    df_desativadas = df_full[mask_aba_full & mask_inativa].copy()

    print("Processando (padronização e regras)...")
    df_ativas = legacy_app.apply_team_standardization(df_ativas, rules_path)
    df_desativadas = legacy_app.apply_team_standardization(df_desativadas, rules_path)
    tbl_ativas = legacy_app.build_full_table(df_ativas)
    tbl_desativadas = legacy_app.build_full_table(df_desativadas)

    # Gerentes do Alimento (apenas em segmento Alimento)
    gerentes_ali = legacy_app._get_gerentes_alimento(rules_path)
    for df_tbl in [tbl_ativas, tbl_desativadas]:
        mask_seg = df_tbl["Segmento"].fillna("").astype(str).str.strip().str.lower() == "alimento"
        mask_ger = (df_tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente") | df_tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_ali
        )
        if (mask_seg & mask_ger).any():
            df_tbl.loc[mask_seg & mask_ger, "EquipePadrao"] = "Gerentes do Alimento"
    # Gerentes do Medicamento (apenas em segmento Medicamento)
    gerentes_med = legacy_app._get_gerentes_medicamento(rules_path)
    for df_tbl in [tbl_ativas, tbl_desativadas]:
        mask_seg = df_tbl["Segmento"].fillna("").astype(str).str.strip().str.lower() == "medicamento"
        mask_ger = (df_tbl["Papel"].fillna("").astype(str).str.strip().str.lower() == "gerente") | df_tbl["Nome"].fillna("").astype(str).apply(
            lambda x: normalize_text(x.strip()) in gerentes_med
        )
        if (mask_seg & mask_ger).any():
            df_tbl.loc[mask_seg & mask_ger, "EquipePadrao"] = "Gerentes do Medicamento"

    print("Salvando no banco...")
    n_ativas = save_linhas(tbl_ativas, modo="ativas")
    n_des = save_linhas(tbl_desativadas, modo="desativadas")
    save_relacao_ativas(ativas_set)

    print(f"OK Linhas ativas: {n_ativas}")
    print(f"OK Linhas desativadas: {n_des}")
    print(f"Banco: {get_db_path()}")


if __name__ == "__main__":
    main()
